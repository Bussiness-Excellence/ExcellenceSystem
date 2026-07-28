#!/usr/bin/env python3
"""
pulpoplus_hierarchy.py
──────────────────────
Parses the 2026 Promotional Team Structure Excel and provides:

  1. load_hierarchy(path)        → HierarchyMap object
  2. HierarchyMap.assign_team()  → resolve team name from employee name or code
  3. HierarchyMap.enrich_records() → stamp Team on Raw-Data records
  4. write_hierarchy_sheet()     → write a formatted "Team Hierarchy" sheet
                                   into an existing openpyxl Workbook

Structure of each sheet in the Excel:
  Row 1 : TEAM   <team_name>
  Row 2 : BLM    <blm_name>   <blm_code>
  Row 3 : header (Division Name / Employee / Employee code)
  Row 4+: data rows — either:
    • MR row   : (seq_number, territory, employee_name, employee_code)
    • Role row : (None,       role_label, employee_name, employee_code)
                  role_label contains "Supervisor" | "Area Manager" | "BUM"
"""

import re
import difflib
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── colour palette (matches existing script style) ───────────────────────────
_C = {
    "blm":        "1F4E79",   # dark navy  – BLM row
    "area_mgr":   "2E75B6",   # medium blue – Area Manager
    "supervisor": "9DC3E6",   # light blue  – Supervisor
    "mr":         "DEEAF1",   # very light  – MR / field rep
    "vacant":     "F2F2F2",   # grey        – vacant slot
    "header":     "1F4E79",   # sheet column headers
    "white":      "FFFFFF",
    "dark_text":  "1A1A2E",
}
_SIDE  = Side(style="thin", color="BDD7EE")
_BORD  = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
_THICK = Side(style="medium", color="1F4E79")


def _norm(name):
    """Normalise a name string for fuzzy matching."""
    if not name:
        return ""
    return re.sub(r"\s+", " ", str(name).strip().lower().replace("\xa0", " "))


# ── manual overrides ──────────────────────────────────────────────────────────
# Add entries here for users whose PulpoPlus name doesn't match the structure
# file. Format: "pulpoplus name (lowercase)" → ("Employee Code", "Team Name")
# These are checked first before any fuzzy matching.
MANUAL_OVERRIDES: dict[str, tuple[str, str]] = {
    # Users confirmed in visits data but absent from the 2026 structure file
    "ahmed abdelmonam nemes":  ("", "FOCUS 1"),   # code unknown — add when available
    "amr aboelhamd":           ("", "FOCUS 1"),   # code unknown — add when available
    "mohamed elrawey":         ("", "FOCUS 1"),   # code unknown — add when available
}


def _is_vacant(val):
    if val is None:
        return True
    return _norm(str(val)) in ("vacant", "")


# ── data model ────────────────────────────────────────────────────────────────

class Employee:
    __slots__ = ("name", "code", "role", "territory", "team", "blm",
                 "area_manager", "supervisor")

    def __init__(self, name, code, role, territory, team, blm,
                 area_manager="", supervisor=""):
        self.name         = str(name or "").strip().replace("\xa0", " ")
        self.code         = str(int(code)) if code and str(code).replace(".0","").isdigit() else ""
        self.role         = role          # "MR" | "Supervisor" | "Area Manager" | "BLM"
        self.territory    = territory
        self.team         = team
        self.blm          = blm
        self.area_manager = area_manager
        self.supervisor   = supervisor


class HierarchyMap:
    """Lookup table built from the 2026 team-structure workbook."""

    def __init__(self):
        self._by_code: dict[str, Employee]   = {}   # employee_code → Employee
        self._by_name: dict[str, Employee]   = {}   # normalised_name → Employee
        self._teams:   dict[str, list[Employee]] = {}  # team_name → [Employee]
        self.all_employees: list[Employee]   = []

    def _add(self, emp: Employee):
        if not emp.name or _is_vacant(emp.name):
            return
        self.all_employees.append(emp)
        if emp.team not in self._teams:
            self._teams[emp.team] = []
        self._teams[emp.team].append(emp)
        if emp.code:
            if emp.code in self._by_code:
                existing = self._by_code[emp.code]
                existing_teams = [t.strip() for t in existing.team.split(";")]
                if emp.team not in existing_teams:
                    existing.team = existing.team + "; " + emp.team
            else:
                self._by_code[emp.code] = emp
        key = _norm(emp.name)
        if key:
            if key in self._by_name:
                existing = self._by_name[key]
                existing_teams = [t.strip() for t in existing.team.split(";")]
                if emp.team not in existing_teams:
                    existing.team = existing.team + "; " + emp.team
            else:
                self._by_name[key] = emp

    # ── public lookup ─────────────────────────────────────────────────────────

    def lookup(self, name=None, code=None) -> "Employee | None":
        """Return the Employee record for a given name or code, or None."""
        if code:
            c = str(code).replace(".0", "").strip()
            if c in self._by_code:
                return self._by_code[c]
        if name:
            key = _norm(name)

            # 0. Manual override table — checked before any fuzzy logic
            if key in MANUAL_OVERRIDES:
                emp_code, emp_team = MANUAL_OVERRIDES[key]
                return Employee(name=name, code=emp_code, role="MR",
                                territory="", team=emp_team, blm="")

            # 1. Exact normalised match
            if key in self._by_name:
                return self._by_name[key]
            # 2. Substring match (stored name contains the query or vice versa)
            for stored_key, emp in self._by_name.items():
                if key and (key in stored_key or stored_key in key):
                    return emp
            # 3. Token overlap: ALL query tokens (>3 chars) found in stored name
            tokens = [t for t in key.split() if len(t) > 3]
            if len(tokens) >= 2:
                for stored_key, emp in self._by_name.items():
                    if all(t in stored_key for t in tokens):
                        return emp
            # 4. Last-token (family name) match — only when the first token also
            #    matches, to avoid false positives like Mohamed Ismaiel → Amr Ismaiel
            if len(tokens) >= 2:
                first, last = tokens[0], tokens[-1]
                hits = [emp for sk, emp in self._by_name.items()
                        if last in sk and first in sk]
                if len(hits) == 1:
                    return hits[0]
            # 5. Similarity fallback — catches minor spelling/transliteration
            #    variants that substring/token matching can't (e.g. "Waled"
            #    vs "Waleed", a single inserted letter, common with Arabic
            #    names transliterated inconsistently between systems).
            #    Requires both high similarity AND a clear margin over the
            #    next-best candidate, so it only fires when there's one
            #    unambiguous near-match rather than guessing among several.
            best_match, best_ratio, runner_up_ratio = None, 0.0, 0.0
            for stored_key, emp in self._by_name.items():
                ratio = difflib.SequenceMatcher(None, key, stored_key).ratio()
                if ratio > best_ratio:
                    runner_up_ratio = best_ratio
                    best_ratio = ratio
                    best_match = emp
                elif ratio > runner_up_ratio:
                    runner_up_ratio = ratio
            if (best_match and best_ratio >= 0.88
                    and (best_ratio - runner_up_ratio) >= 0.05):
                return best_match
        return None

    def assign_team(self, name=None, code=None) -> str:
        """Return the team name string, or '' if not found."""
        emp = self.lookup(name=name, code=code)
        return emp.team if emp else ""

    # ── record enrichment ─────────────────────────────────────────────────────

    def enrich_records(self, records: list[dict], warn: bool = True) -> int:
        """
        Stamp the `team` field on every record using the hierarchy (only if
        currently blank), and stamp a separate `hierarchy_territory` field
        with the employee's official territory from the hierarchy file.

        `hierarchy_territory` is kept apart from the visit-derived
        `territory` field on purpose: `territory` reflects which account a
        given visit was actually logged against, and downstream logic
        (manager detection, territory backfill) depends on that per-visit
        signal. Overwriting it with one fixed hierarchy value per person
        would silently break manager detection, which relies on a rep
        visiting more than one distinct territory. `hierarchy_territory`
        is used only when building the displayed Territory column.

        Returns the count of records that were successfully enriched
        (team and/or hierarchy_territory).
        """
        enriched = 0
        unmatched: set[str] = set()

        for r in records:
            r.setdefault("hierarchy_territory", "")
            emp = self.lookup(name=r.get("user"), code=None)
            if not emp:
                if not str(r.get("team") or "").strip():
                    unmatched.add(r.get("user", ""))
                continue

            did_enrich = False
            if not str(r.get("team") or "").strip():
                r["team"] = emp.team
                did_enrich = True

            if emp.territory and str(emp.territory).strip():
                r["hierarchy_territory"] = emp.territory
                did_enrich = True

            if did_enrich:
                enriched += 1

        if warn and unmatched:
            print(f"  ⚠️  {len(unmatched)} user(s) not found in hierarchy "
                  f"(team left blank): {sorted(unmatched)[:10]}"
                  + (" …" if len(unmatched) > 10 else ""))
        return enriched

    def teams(self):
        return list(self._teams.keys())

    def members_of(self, team):
        return self._teams.get(team, [])


# ── parser ────────────────────────────────────────────────────────────────────

def _detect_team_and_blm(ws):
    """
    Read rows 1-3 of a sheet to find the team name and BLM.
    Handles the inconsistent column layouts across sheets.
    """
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 3:
            break
        rows.append(row)

    team_name = ""
    blm_name  = ""
    blm_code  = ""

    for row in rows:
        for j, cell in enumerate(row):
            if cell is None:
                continue
            s = str(cell).strip().replace("\xa0", " ")

            # BUM / BLM inline format: "BUM: Ahmed Hussien" in one cell,
            # with the numeric code in the next cell to the right
            bum_m = re.match(r"(?:BUM|BLM)\s*[:\s]+(.+)", s, re.I)
            if bum_m:
                blm_name = bum_m.group(1).strip()
                # code is in the next non-None cell to the right
                for k in range(j + 1, min(j + 4, len(row))):
                    if row[k] is not None:
                        try:
                            blm_code = str(int(float(str(row[k]))))
                        except Exception:
                            pass
                        break
                continue

            # Standard "Business Line Manager(BLM)" label row
            if re.search(r"Business Line", s, re.I):
                for k in range(j + 1, min(j + 4, len(row))):
                    if row[k]:
                        blm_name = str(row[k]).strip()
                        if k + 1 < len(row) and row[k + 1]:
                            try:
                                blm_code = str(int(float(str(row[k + 1]))))
                            except Exception:
                                pass
                        break
                continue

            # TEAM keyword in col B, team name in col C
            if s.upper() == "TEAM" and j == 1 and len(row) > 2:
                val = row[2]
                if val:
                    candidate = str(val).strip().replace("\xa0", " ")
                    # Reject if it looks like a person's name without team keywords
                    # but accept it for now; swap logic below handles the rest
                    if candidate and not re.match(r"\d+", candidate):
                        team_name = candidate
                continue

            # TEAM keyword embedded: "TEAM : EAGLES 1"
            if re.search(r"TEAM\s*[:\s]", s, re.I):
                m = re.search(r"TEAM\s*[:\s]+(.+)", s, re.I)
                if m:
                    candidate = m.group(1).strip()
                    if candidate and not re.match(r"(BLM|BUM|Manager)", candidate, re.I):
                        team_name = candidate

    # Fallback: use the sheet name itself as team name
    if not team_name:
        team_name = ws.title

    # Clean up team name
    team_name = re.sub(r"\s+", " ", team_name).strip()

    # Sanity check: if team_name looks like a person's name (no team keywords)
    # and blm_name looks like a team keyword — they are swapped (FALCONS layout).
    # Also grab the code from col D row 0 for FALCONS (it's in the same row as the name).
    team_keywords = r"(PLATINUM|STAR|FOCUS|PIONEER|EAGLE|FALCON)"
    if (not re.search(team_keywords, team_name, re.I)
            and re.search(team_keywords, blm_name, re.I)):
        team_name, blm_name = blm_name, team_name
        # For FALCONS the code is in col D of row 0 alongside the BLM name
        if not blm_code and len(rows) > 0:
            r0 = rows[0]
            for cell in r0:
                if cell and str(cell).strip().isdigit():
                    blm_code = str(int(cell))
                    break

    return team_name, blm_name, blm_code


def _role_from_label(label: str) -> str:
    label = str(label or "").strip()
    lo = label.lower()
    if "area manager" in lo:
        return "Area Manager"
    if "supervisor" in lo:
        return "Supervisor"
    if re.search(r"blm|bum|business line", lo):
        return "BLM"
    return "Unknown"


_FLAT_HEADER = ["Team", "BLM", "Area Manager", "Supervisor", "Territory",
                "Employee Name", "Employee Code", "Role"]


def _looks_like_flat_hierarchy(wb) -> bool:
    """
    Detect the already-flattened 'Team Hierarchy' sheet this module itself
    writes via write_hierarchy_sheet() (e.g. if a user re-exported that tab
    as its own file). Its header row is fixed and unmistakable, so this is
    a safe, exact check rather than a fuzzy heuristic.
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        first_row = next(ws.iter_rows(values_only=True), None)
        if first_row and list(first_row[:len(_FLAT_HEADER)]) == _FLAT_HEADER:
            return True
    return False


def _load_flat_hierarchy(wb) -> HierarchyMap:
    """
    Load a HierarchyMap from the flat 'Team Hierarchy' output format
    (Team, BLM, Area Manager, Supervisor, Territory, Employee Name,
    Employee Code, Role) instead of the raw multi-sheet team-structure
    workbook. Every row already carries a resolved code, so no fuzzy
    name-matching is needed here — this is a straight, literal load.
    """
    hmap = HierarchyMap()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header or list(header[:len(_FLAT_HEADER)]) != _FLAT_HEADER:
            continue  # not the flat sheet (e.g. a stray extra tab) — skip it

        for row in rows:
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue  # blank separator row before the summary block
            team, blm, area_mgr, supervisor, territory, name, code, role = (
                row + (None,) * (len(_FLAT_HEADER) - len(row))
            )[:len(_FLAT_HEADER)]
            if not role or str(team or "").strip().lower() == "total people in hierarchy" \
                    or str(team or "").strip().lower() == "teams":
                break  # reached the trailing summary block — stop this sheet
            if _is_vacant(name):
                continue
            hmap._add(Employee(
                name=name, code=code, role=str(role or "").strip(),
                territory=territory or "", team=str(team or "").strip(),
                blm=blm or "", area_manager=area_mgr or "",
                supervisor=supervisor or "",
            ))

    print(f"  ✓ Hierarchy loaded from flat 'Team Hierarchy' sheet: "
          f"{len(hmap.all_employees)} people across {len(hmap.teams())} teams")
    return hmap


def load_hierarchy(path: str) -> HierarchyMap:
    """
    Parse the team-structure workbook and return a HierarchyMap.
    Auto-detects whether it's the raw multi-sheet source structure file or
    an already-flattened 'Team Hierarchy' export, and parses accordingly.
    """
    wb = load_workbook(path, read_only=True, data_only=True)

    if _looks_like_flat_hierarchy(wb):
        hmap = _load_flat_hierarchy(wb)
        wb.close()
        return hmap

    hmap = HierarchyMap()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        team_name, blm_name, blm_code = _detect_team_and_blm(ws)

        # Add the BLM itself
        if blm_name and not _is_vacant(blm_name):
            hmap._add(Employee(
                name=blm_name, code=blm_code, role="BLM",
                territory="", team=team_name, blm=blm_name,
            ))

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue

        # Find all Area Manager header rows and their row indexes
        am_rows = []
        for r_idx, row in enumerate(rows[3:], 4):
            seq = row[0]
            col_b = str(row[1] or "").strip().replace("\xa0", " ")
            col_c = str(row[2] or "").strip().replace("\xa0", " ")
            col_d = row[3]
            if seq is None and _role_from_label(col_b) == "Area Manager":
                am_rows.append((r_idx, col_c, col_d, col_b))

        start_row = 4
        current_area_mgr = ""

        for am_r_idx, am_name, am_code, am_terr in am_rows:
            current_area_mgr = am_name
            if am_name and not _is_vacant(am_name):
                hmap._add(Employee(
                    name=am_name, code=am_code, role="Area Manager",
                    territory=am_terr, team=team_name, blm=blm_name,
                    area_manager=am_name,
                ))

            current_supervisor = ""

            for r_idx in range(start_row, am_r_idx):
                row = rows[r_idx - 1]
                seq = row[0]
                col_b = str(row[1] or "").strip().replace("\xa0", " ")
                col_c = str(row[2] or "").strip().replace("\xa0", " ")
                col_d = row[3]
                parent_sup = str(row[4] or "").strip().replace("\xa0", " ") if len(row) > 4 and row[4] else ""

                role = _role_from_label(col_b) if seq is None else "MR"

                if seq is None and role == "Supervisor":
                    current_supervisor = col_c
                    if col_c and not _is_vacant(col_c):
                        hmap._add(Employee(
                            name=col_c, code=col_d, role=role,
                            territory=col_b, team=team_name, blm=blm_name,
                            area_manager=current_area_mgr,
                            supervisor=col_c,
                        ))
                    continue

                if seq is not None:
                    name = col_c
                    code = str(col_d).strip() if col_d is not None else ""
                    territory = col_b
                    if not name or _is_vacant(name):
                        continue
                    hmap._add(Employee(
                        name=name, code=code, role="MR",
                        territory=territory, team=team_name, blm=blm_name,
                        area_manager=current_area_mgr,
                        supervisor=parent_sup or current_supervisor,
                    ))

            start_row = am_r_idx + 1

        # Fallback for remaining rows after the last AM block if any
        if start_row <= len(rows):
            current_supervisor = ""
            for r_idx in range(start_row, len(rows) + 1):
                row = rows[r_idx - 1]
                seq = row[0]
                col_b = str(row[1] or "").strip().replace("\xa0", " ")
                col_c = str(row[2] or "").strip().replace("\xa0", " ")
                col_d = row[3]
                parent_sup = str(row[4] or "").strip().replace("\xa0", " ") if len(row) > 4 and row[4] else ""

                role = _role_from_label(col_b) if seq is None else "MR"
                if seq is None and role == "Supervisor":
                    current_supervisor = col_c
                    if col_c and not _is_vacant(col_c):
                        hmap._add(Employee(
                            name=col_c, code=col_d, role=role,
                            territory=col_b, team=team_name, blm=blm_name,
                            area_manager=current_area_mgr,
                            supervisor=col_c,
                        ))
                    continue
                if seq is not None:
                    name = col_c
                    code = str(col_d).strip() if col_d is not None else ""
                    territory = col_b
                    if not name or _is_vacant(name):
                        continue
                    hmap._add(Employee(
                        name=name, code=code, role="MR",
                        territory=territory, team=team_name, blm=blm_name,
                        area_manager=current_area_mgr,
                        supervisor=parent_sup or current_supervisor,
                    ))

    wb.close()
    print(f"  ✓ Hierarchy loaded: {len(hmap.all_employees)} people "
          f"across {len(hmap.teams())} teams")
    return hmap


# ── Excel sheet writer ────────────────────────────────────────────────────────

def write_hierarchy_sheet(wb, hmap: HierarchyMap,
                          sheet_name: str = "Team Hierarchy",
                          position: int = 0):
    """
    Write a formatted 'Team Hierarchy' sheet into an existing openpyxl Workbook.
    Inserts the sheet at `position` (default: first tab).
    """
    # Remove existing sheet with the same name if present
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name, position)

    # ── column layout ─────────────────────────────────────────────────────────
    COLS = ["Team", "BLM", "Area Manager", "Supervisor", "Territory",
            "Employee Name", "Employee Code", "Role"]
    col_widths = [14, 22, 24, 24, 26, 30, 14, 14]

    # Header row
    for j, (col, w) in enumerate(zip(COLS, col_widths), start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font      = Font(name="Arial", bold=True, color=_C["white"], size=10)
        cell.fill      = PatternFill("solid", fgColor=_C["header"])
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = _BORD
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Role → fill colour mapping
    role_fill = {
        "BLM":          PatternFill("solid", fgColor=_C["blm"]),
        "Area Manager": PatternFill("solid", fgColor=_C["area_mgr"]),
        "Supervisor":   PatternFill("solid", fgColor=_C["supervisor"]),
        "MR":           PatternFill("solid", fgColor=_C["mr"]),
    }
    role_font_color = {
        "BLM":          _C["white"],
        "Area Manager": _C["white"],
        "Supervisor":   _C["dark_text"],
        "MR":           _C["dark_text"],
    }

    # Sort: by team, then role priority, then name
    role_order = {"BLM": 0, "Area Manager": 1, "Supervisor": 2, "MR": 3}
    sorted_emps = sorted(
        hmap.all_employees,
        key=lambda e: (e.team, role_order.get(e.role, 9), e.name)
    )

    for row_idx, emp in enumerate(sorted_emps, start=2):
        fill  = role_fill.get(emp.role, PatternFill("solid", fgColor=_C["mr"]))
        fcolor = role_font_color.get(emp.role, _C["dark_text"])
        font  = Font(name="Arial", size=9, color=fcolor,
                     bold=(emp.role in ("BLM", "Area Manager")))

        values = [emp.team, emp.blm, emp.area_manager, emp.supervisor,
                  emp.territory, emp.name, emp.code, emp.role]

        for j, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=j, value=val or "")
            cell.font      = font
            cell.fill      = fill
            cell.border    = _BORD
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=False)

        ws.row_dimensions[row_idx].height = 16

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # Summary block in the first few rows after the data
    last_data_row = len(sorted_emps) + 1
    summary_row   = last_data_row + 2
    ws.cell(row=summary_row, column=1, value="Total People in Hierarchy").font = \
        Font(name="Arial", bold=True, size=9, color=_C["blm"])
    ws.cell(row=summary_row, column=2, value=len(hmap.all_employees)).font = \
        Font(name="Arial", size=9)

    ws.cell(row=summary_row + 1, column=1, value="Teams").font = \
        Font(name="Arial", bold=True, size=9, color=_C["blm"])
    ws.cell(row=summary_row + 1, column=2,
            value=", ".join(sorted(hmap.teams()))).font = \
        Font(name="Arial", size=9)

    print(f"  ✓ '{sheet_name}' sheet written: {len(sorted_emps)} rows, "
          f"{len(hmap.teams())} teams")
    return ws