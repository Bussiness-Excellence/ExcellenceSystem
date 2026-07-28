#!/usr/bin/env python3
"""
PulpoPlus CRM Data Extractor - SMART SUMMARY VERSION
Instead of dumping every raw table, this:
  1. Auto-detects the real "Visits" report table on the page (by header
     content, not by guessing table position).
  2. Parses every visit row into a clean record: user, date, time,
     account type, doctor, etc.
  3. Aggregates PER USER across the whole date range into KPIs:
       - Number of working days
       - Covered doctors per account type (Clinic / Poly Clinic / AM Center / Hospital)
       - Number of Events
       - Number of Office Work entries
       - Average field duration per day (last visit time - first visit time)
  4. Saves a clean "Summary" sheet + a "Raw Data" sheet to Excel.

Usage:
  python pulpoplus_extractor_summary.py --from 2026-06-01 --to 2026-06-10 --debug
"""

import re
import time
from datetime import datetime, timedelta
from collections import defaultdict

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import argparse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION
# ============================================================
LOGIN_URL = "https://pulpo-eipico.cloud/crm/"
USERNAME = "be"
PASSWORD = "1234"

# Maps the raw code/text found in the "Acc. Type" column to a normalized
# category. IMPORTANT: run once with --debug first -- the script will print
# every unique raw Acc. Type value it actually found on the page ("RAW ACC
# TYPE VALUES SEEN"). Compare that list to the keys below and fix/extend
# this dict if anything shows up as "Other" that shouldn't.
ACC_TYPE_MAP = {
    "C": "Clinic",
    "CLINIC": "Clinic",
    "P": "Pharmacy",
    "PH": "Pharmacy",
    "PHARMACY": "Pharmacy",
    "H": "Hospital",
    "HOSPITAL": "Hospital",
    "AM": "AM Center",
    "AM CENTER": "AM Center",
    "PC": "Poly Clinics",
    "POLY CLINIC": "Poly Clinics",
    "POLY CLINICS": "Poly Clinics",
    "D": "Distributors",
    "DISTRIBUTOR": "Distributors",
    "DISTRIBUTORS": "Distributors",
    "OW": "Office Work",
    "OFFICE WORK": "Office Work",
    "A": "Activities",
    "ACTIVITY": "Activities",
    "ACTIVITIES": "Activities",
    "E": "Events",
    "EVENT": "Events",
    "EVENTS": "Events",
}

# The 4 categories the person asked to break "covered doctors" down by.
COVERAGE_ACC_TYPES = ["Clinic", "Poly Clinics", "AM Center", "Hospital"]

# Flat set of every raw code AND every mapped label from ACC_TYPE_MAP,
# lowercased. Used to detect the colspan-duplication artifact where a
# merged Office Work / Events row echoes its Acc. Type text into the
# Territory cell too (see the guard in build_records below) -- checking
# against this full set (rather than just the current row's own raw text)
# means the guard fires regardless of whether the page happened to render
# the short code ("OW") or the full label ("Office Work") that time.
ACC_TYPE_LABELS_LOWER = {k.lower() for k in ACC_TYPE_MAP.keys()} | {v.lower() for v in ACC_TYPE_MAP.values()}

# --- Acc. Type CHECKBOXES on the search form (Clinic / Pharmacy / Hospital /
# AM Center / Poly Clinics / Distributors / Office Work / Activities / Events) ---
# The site remembers whatever was last checked (cookie/session), which can
# silently make "Show" return the WRONG account types (e.g. Pharmacy-only)
# without any error. To avoid this, we explicitly force these checkboxes to
# an exact known state before every "Show" click: ONLY the types relevant
# to our doctor-coverage / AM-PM analysis are checked, everything else is
# unchecked. Match is case-insensitive and whitespace-normalized.
# "Office Work" and "Events" are included even though they're not doctor
# calls: Office Work can stand in for an AM field shift (see AM Slot Days /
# Total Shifts below), and Events feeds "No. of Events". Both were
# previously EXCLUDED from the search filter entirely, which is why those
# two counts were always 0 no matter what was on the page.
ACC_TYPE_INCLUDE_LABELS = {"clinic", "hospital", "am center", "poly clinics", "office work", "events", "event"}

# --- AM / PM shift grouping ---
# Per the user's rule: Clinic + Poly Clinics = "PM" visits.
#                       Hospital + AM Center  = "AM" visits.
# Anything else (Pharmacy, Distributors, Office Work, Activities, Events,
# Other) is not a doctor "call" and is excluded from AM/PM call/coverage/
# specialty math (it's still counted separately via No. of Events / No. of
# Office Work).
SHIFT_MAP = {
    "Hospital": "AM",
    "AM Center": "AM",
    "Clinic": "PM",
    "Poly Clinics": "PM",
}

# Maps the raw code/text found in the "Visit Type" column (e.g. 'S' / 'D')
# to Single / Double. IMPORTANT: run with --debug first and check the
# "RAW VISIT TYPE VALUES SEEN" line -- fix/extend this if codes differ.
VISIT_TYPE_MAP = {
    "S": "Single",
    "SINGLE": "Single",
    "D": "Double",
    "DOUBLE": "Double",
}

# Column name synonyms -- the report's exact header text may vary slightly
# (e.g. "Doc. ID" vs "Doctor ID"), so we match loosely by keyword.
COLUMN_KEYWORDS = {
    "member": ["members", "member"],
    "acc_type": ["acc. type", "acc type", "account type"],
    "acc_name": ["acc. name", "acc name", "account name"],
    "doc_id": ["doc. id", "doc id", "doctor id"],
    "acc_id": ["acc. id", "acc id", "account id"],
    "doctor_name": ["doctor name"],
    "date": ["date"],
    "territory": ["territory"],
    "visit_type": ["visit type"],
    "specialty": ["spec.", "spec", "specialty"],
    "classification": ["class"],
    "products": ["products", "product"],
    # Notes: free-text field where the rep writes the exact doctor name.
    # When Doctor Name + Specialty + Date are identical for two rows, the Notes
    # value is what proves they are different customers.  Including notes in
    # doctor_key prevents those rows being collapsed into one unique doctor.
    "notes": ["notes", "note"],
}


def parse_arguments():
    parser = argparse.ArgumentParser(description="Smart per-user KPI summary from PulpoPlus visit reports")
    parser.add_argument("--from", dest="from_date", required=False, help="Start date (YYYY-MM-DD). Default: 14 days ago")
    parser.add_argument("--to", dest="to_date", required=False, help="End date (YYYY-MM-DD). Default: today")
    parser.add_argument("--debug", action="store_true", help="Enable debug output")
    parser.add_argument("--max-wait", dest="max_wait", type=int, default=180, help="Max wait time for report (seconds)")
    parser.add_argument("--settle-time", dest="settle_time", type=int, default=8,
                         help="Extra seconds to wait AFTER page looks stable (default 8s)")
    parser.add_argument("--max-rows-per-table", dest="max_rows_per_table", type=int, default=20000,
                         help="Tables with more rows than this AND no identifiable fieldset/legend "
                              "section title are skipped as junk widget tables (default 20000). "
                              "Named tables like 'Visits' are never skipped by this, regardless of size.")
    parser.add_argument("--max-cols-per-table", dest="max_cols_per_table", type=int, default=200,
                         help="Tables with more columns than this are skipped as junk layout tables (default 200)")
    parser.add_argument("--team", dest="team", default=None,
                         help="Run against a SINGLE team only, for a quick test. Match is case-insensitive "
                              "and can be a partial name (e.g. --team \"FOCUS 2\" or --team focus). "
                              "If omitted, all teams are processed.")
    parser.add_argument("--list-teams", action="store_true",
                         help="Just print all available team names and exit (no extraction).")
    return parser.parse_args()


def get_date_range(from_date_str, to_date_str):
    if to_date_str:
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d")
    else:
        to_date = datetime.now()

    if from_date_str:
        from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
    else:
        from_date = to_date - timedelta(days=14)

    return from_date.strftime("%Y-%m-%d"), to_date.strftime("%Y-%m-%d")


# ============================================================
# BROWSER SETUP / LOGIN / NAVIGATION
# ============================================================

def setup_driver():
    print("🚀 Starting browser...")

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    # IMPORTANT: this is a DIFFERENT timeout than set_page_load_timeout /
    # set_script_timeout below. Those control how long the BROWSER is
    # allowed to keep running a page load / injected script before IT
    # gives up. This controls how long Selenium's own local HTTP client
    # (talking to chromedriver over localhost) will wait to READ a
    # response before giving up on ITS end.
    #
    # With large pulls (10k-15k+ visit rows), the EXTRACT_ALL_TABLES_JS
    # script finishes fine inside the browser, but chromedriver can take
    # well over 120s to serialize and ship that huge JSON payload back
    # over the wire. Selenium's default read timeout for that local
    # connection is 120s, which is what produced:
    #   HTTPConnectionPool(host='localhost', port=52803): Read timed out (read timeout=120)
    # Raising set_script_timeout() alone does NOT fix this -- it's a
    # different timeout entirely.
    #
    # NOTE: RemoteConnection.set_timeout(...) called as a CLASSMETHOD
    # before any connection instance exists is broken/deprecated in some
    # selenium builds (raises AttributeError on _client_config), so we set
    # it on the live instance instead, after the driver is up, trying each
    # known selenium API in turn and ignoring whichever don't apply.
    new_timeout = 400
    executor = driver.command_executor
    for attempt in (
        lambda: setattr(executor._client_config, "timeout", new_timeout),
        lambda: executor.set_timeout(new_timeout),
        lambda: setattr(executor, "_timeout", new_timeout),
    ):
        try:
            attempt()
            break
        except Exception:
            continue

    driver.set_page_load_timeout(300)
    driver.set_script_timeout(300)
    print("   ✓ Browser started")
    return driver


def login(driver, max_attempts=3):
    print("🔐 Logging in...")

    # driver.get() can fail with "Timed out receiving message from
    # renderer: 120.000" if the site is just slow/flaky that moment --
    # this is a transient network/site hiccup, not a code bug, so it's
    # worth a couple of retries before giving up on the whole run.
    for attempt in range(1, max_attempts + 1):
        try:
            driver.get(LOGIN_URL)
            break
        except Exception as e:
            print(f"   ⚠️  Page load attempt {attempt}/{max_attempts} failed: {e}")
            if attempt == max_attempts:
                print("❌ Login failed: could not load login page after retries")
                return False
            time.sleep(5)

    time.sleep(2)

    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located((By.TAG_NAME, "input"))
        )
        time.sleep(2)

        inputs = driver.find_elements(By.TAG_NAME, "input")
        if len(inputs) >= 2:
            inputs[0].send_keys(USERNAME)
            inputs[1].send_keys(PASSWORD)
            inputs[1].send_keys(Keys.RETURN)

        time.sleep(5)
        print("✓ Login successful")
        return True
    except Exception as e:
        print(f"❌ Login failed: {e}")
        return False


def navigate_to_reports(driver):
    print("📊 Loading reports page...")
    report_url = "https://pulpo-eipico.cloud/crm/report_ter_visits_search.php"
    driver.get(report_url)
    time.sleep(3)

    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "team"))
        )

        team_dropdown = driver.find_element(By.NAME, "team")
        team_options = team_dropdown.find_elements(By.TAG_NAME, "option")
        teams = []

        for option in team_options:
            value = option.get_attribute("value")
            text = option.text.strip()
            if value and text and text.lower() not in ['select team', 'please select']:
                teams.append({"value": value, "name": text})

        print(f"✓ Found {len(teams)} teams")
        return teams

    except Exception as e:
        print(f"❌ Error: {e}")
        return []


def set_date_field(driver, field_name, date_string):
    try:
        script = f"""
        var field = document.querySelector('input[name="{field_name}"]');
        if (field) {{
            field.value = "{date_string}";
            field.dispatchEvent(new Event('change', {{ bubbles: true }}));
            return "OK";
        }}
        return "FAIL";
        """
        return driver.execute_script(script) == "OK"
    except Exception:
        return False


def click_show_button(driver, max_attempts=3):
    print("   🔘 Clicking Show button...")
    try:
        try:
            btn = driver.find_element(By.CSS_SELECTOR, 'input[type="button"][value="Show"]')
        except Exception:
            btn = driver.find_element(By.XPATH, "//input[@value='Show']")
    except Exception as e:
        print(f"      ❌ Click failed: could not locate Show button: {e}")
        return False

    # Clicking Show triggers the report to render thousands of rows into
    # the DOM, which can pin the renderer thread busy long enough to trip
    # "Timed out receiving message from renderer" -- the same transient
    # failure mode as the login page load, just triggered by a click
    # instead of driver.get(). Retry rather than aborting the whole run.
    for attempt in range(1, max_attempts + 1):
        try:
            driver.execute_script("arguments[0].click();", btn)
            print("      ✓ Click executed")
            return True
        except Exception as e:
            print(f"      ⚠️  Click attempt {attempt}/{max_attempts} failed: {e}")
            if attempt == max_attempts:
                print("      ❌ Click failed after retries")
                return False
            time.sleep(5)


GET_CHECKBOX_LABELS_JS = r"""
function getLabelForCheckbox(cb) {
    // 1) explicit <label for="id">
    if (cb.id) {
        const lbl = document.querySelector('label[for="' + cb.id + '"]');
        if (lbl) return lbl.textContent.trim();
    }
    // 2) checkbox wrapped inside a <label>
    const wrapping = cb.closest('label');
    if (wrapping) return wrapping.textContent.trim();
    // 3) fallback: text of following sibling node(s) up to the next <input>
    let text = '';
    let node = cb.nextSibling;
    while (node) {
        if (node.nodeType === 1 && node.tagName === 'INPUT') break;
        text += node.textContent || '';
        node = node.nextSibling;
    }
    return text.trim();
}
const checkboxes = Array.from(document.querySelectorAll('input[type="checkbox"]'));
return checkboxes.map(cb => ({
    name: cb.name, value: cb.value, checked: cb.checked, label: getLabelForCheckbox(cb)
}));
"""


def set_acc_type_filters(driver, debug=False):
    """
    Forces the Acc. Type checkboxes on the search form to an EXACT known
    state -- only Clinic, Poly Clinics, AM Center, Hospital checked -- no
    matter what the page/session had checked by default. This is critical:
    a prior run silently pulled 100% Pharmacy-only data because the site
    remembered a stale checkbox state from a previous manual session, and
    nothing in the page indicated an error -- the report just returned
    "successfully" with the wrong account types.
    """
    try:
        checkbox_info = driver.execute_script(GET_CHECKBOX_LABELS_JS)
    except Exception as e:
        print(f"   ⚠️  Could not read Acc. Type checkboxes: {e}")
        return False

    elements = driver.find_elements(By.CSS_SELECTOR, 'input[type="checkbox"]')

    if len(elements) != len(checkbox_info):
        print(f"   ⚠️  Checkbox count mismatch ({len(elements)} elements vs "
              f"{len(checkbox_info)} labels read) -- skipping auto-filter, verify manually")
        return False

    changed = 0
    matched_any = False
    for el, info in zip(elements, checkbox_info):
        label_norm = normalize_header(info.get("label", ""))
        if not label_norm:
            continue  # not one of the labeled Acc. Type checkboxes
        should_be_checked = label_norm in ACC_TYPE_INCLUDE_LABELS
        is_checked = bool(info.get("checked"))
        if should_be_checked:
            matched_any = True
        if should_be_checked != is_checked:
            try:
                driver.execute_script("arguments[0].click();", el)
                changed += 1
                if debug:
                    action = "checked" if should_be_checked else "unchecked"
                    print(f"      ✓ {action}: {info['label']!r}")
            except Exception as e:
                print(f"   ⚠️  Could not toggle checkbox {info['label']!r}: {e}")

    if not matched_any:
        print("   ⚠️  None of the Acc. Type checkbox labels matched "
              f"{sorted(ACC_TYPE_INCLUDE_LABELS)} -- verify the labels on the page match expectations")
        return False

    print(f"   ✓ Acc. Type filter forced to Clinic + Poly Clinics + AM Center + Hospital + Office Work + Events "
          f"({changed} checkbox(es) changed)")
    return True


def wait_for_full_page_load(driver, max_wait=180, settle_time=8):
    print(f"   ⏳ Waiting for FULL page load (up to {max_wait}s)...")
    start = time.time()

    try:
        WebDriverWait(driver, max_wait).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        print(f"      ✓ document.readyState = complete ({int(time.time()-start)}s)")
    except Exception:
        print("      ⚠️  Timed out waiting for readyState complete, continuing anyway")

    try:
        WebDriverWait(driver, max_wait).until(
            lambda d: d.execute_script(
                "return (typeof jQuery === 'undefined') || jQuery.active === 0"
            )
        )
        print(f"      ✓ No pending AJAX ({int(time.time()-start)}s)")
    except Exception:
        print("      ⚠️  jQuery idle check timed out/not applicable, continuing anyway")

    stable_checks_needed = 3
    stable_count = 0
    last_row_count = -1

    while time.time() - start < max_wait:
        try:
            total_rows = driver.execute_script(
                "return document.querySelectorAll('table tr').length;"
            )
        except Exception:
            total_rows = -1

        if total_rows == last_row_count and total_rows > 0:
            stable_count += 1
        else:
            stable_count = 0

        last_row_count = total_rows

        if stable_count >= stable_checks_needed:
            elapsed = int(time.time() - start)
            print(f"      ✓ Row count stable at {total_rows} rows ({elapsed}s)")
            break

        elapsed = int(time.time() - start)
        if elapsed % 10 == 0:
            print(f"      ⏳ Still loading... rows so far: {total_rows} ({elapsed}s)")

        time.sleep(1)
    else:
        print(f"      ⚠️  Max wait reached ({max_wait}s), proceeding with whatever loaded")

    print(f"      💤 Settling for an extra {settle_time}s...")
    time.sleep(settle_time)
    return True


# ============================================================
# BULK TABLE EXTRACTION (fast, single JS round-trip)
# ============================================================

EXTRACT_ALL_TABLES_JS = r"""
function cellText(cell) {
    return (cell.innerText || cell.textContent || "").trim();
}

function extractTable(table) {
    let headers = [];
    let bodyRows = [];

    const theadThs = table.querySelectorAll("thead th");
    if (theadThs.length > 0) {
        // Expand each header by its colSpan too, in case a header cell
        // itself visually spans multiple columns.
        headers = [];
        Array.from(theadThs).forEach(h => {
            const span = h.colSpan || 1;
            const text = cellText(h);
            for (let s = 0; s < span; s++) headers.push(text);
        });
        bodyRows = Array.from(table.querySelectorAll("tbody tr"));
        if (bodyRows.length === 0) {
            const allRows = Array.from(table.querySelectorAll("tr"));
            bodyRows = allRows.slice(1);
        }
    } else {
        const allRows = Array.from(table.querySelectorAll("tr"));
        if (allRows.length > 0) {
            const firstRowThs = allRows[0].querySelectorAll("th");
            if (firstRowThs.length > 0) {
                headers = [];
                Array.from(firstRowThs).forEach(h => {
                    const span = h.colSpan || 1;
                    const text = cellText(h);
                    for (let s = 0; s < span; s++) headers.push(text);
                });
                bodyRows = allRows.slice(1);
            } else {
                const firstRowTds = allRows[0].querySelectorAll("td");
                let colCount = 0;
                Array.from(firstRowTds).forEach(td => colCount += (td.colSpan || 1));
                headers = [];
                for (let i = 0; i < colCount; i++) headers.push("Column " + (i + 1));
                bodyRows = allRows;
            }
        }
    }

    // --- Strip a single leading BLANK header cell ---
    // The Visits report table renders a colored priority bar on the far-left
    // of every data row using ONE <th> cell with NO text (empty string). That
    // single blank occupies slot 0 of the headers array, which shifts every
    // subsequent column one position right, so "Members" lands in the wrong
    // slot and every row returns Members='' → user=None → dropped.
    // Fix: if exactly ONE leading header is blank, absorb it into leadingOffset
    // so data cells are mapped starting from the next real header.
    //
    // IMPORTANT: we only do this for EXACTLY 1 leading blank. A layout or
    // navigation table that has a colSpan-expanded blank <th> can produce
    // dozens of consecutive empty strings; stripping those would corrupt
    // the header array for tables we don't care about, and the 99-column
    // warning in the debug log is the symptom of that misfiring.
    let leadingBlankHeaders = 0;
    if (headers.length > 0 && headers[0] === "") {
        // Walk forward only if consecutive blanks at start
        let blanks = 0;
        for (let i = 0; i < headers.length; i++) {
            if (headers[i] === "") { blanks++; } else { break; }
        }
        // Only strip if it's 1-2 blanks (the colored bar pattern).
        // More than 2 = almost certainly a layout table; leave it alone.
        if (blanks <= 2) {
            leadingBlankHeaders = blanks;
            headers = headers.slice(leadingBlankHeaders);
        }
    }

    // --- Detect a leading offset of UNLABELED cells ---
    // NOTE: We previously also tried to auto-detect unlabeled leading
    // columns by comparing headers.length to the MAX effective (colSpan-
    // summed) cell count across a sample of body rows, on the theory that
    // consistently-wider data rows imply unlabeled leading <td>s. In
    // practice this heuristic is unreliable: a single anomalous/malformed
    // row in the 10-row sample (a stray huge colSpan, a promo/legend row,
    // etc.) is enough to produce a wildly wrong offset -- confirmed in
    // production, where it invented 4 fake leading columns on the real
    // 'Visits' table (whose headers already fully and correctly describe
    // every physical column) and shifted every field, and separately
    // invented 246,573(!) fake leading columns on another table from the
    // same page. Both are worse than doing nothing. We now trust ONLY the
    // explicit-blank-leading-<th> detection above, which is driven by
    // actual header markup rather than guessed from row width.
    let leadingOffset = leadingBlankHeaders;

    // --- Build the row grid, honoring BOTH colSpan and rowSpan ---
    // PulpoPlus groups every visit to the SAME account into one visual
    // block by putting Territory/Acc Type/Acc ID/Acc Name into a single
    // cell with rowSpan = (number of visits in that account that day),
    // rendered ONCE at the top of the block. Every subsequent <tr> in
    // that block therefore has FEWER actual <td> elements in the real
    // DOM -- the rowspanned cells simply don't exist on those rows.
    //
    // Mapping cells to headers by raw position (as if every row had the
    // same cell count) silently shifts every field after the rowspan by
    // however many columns were consumed -- Date ends up holding the
    // Members text, Doc ID ends up holding a product code, etc. -- and
    // the resulting garbage "date" fails to parse, so those rows get
    // silently dropped entirely downstream. (Confirmed: this is exactly
    // why every "Double" visit row in a same-account group was going
    // missing while the plain, ungrouped "Single" rows extracted fine.)
    //
    // Fix: reconstruct the true visual grid, the same way a browser
    // renders it. At each column position we either (a) carry forward a
    // still-active rowSpan from an earlier row, or (b) consume the next
    // real <td> from THIS row -- and if that <td> itself has rowSpan > 1,
    // register it to keep carrying forward into the following rows too.
    const rowSpanCarry = {};   // colPosition -> { value, rowsLeft }
    const totalCols = leadingOffset + headers.length;

    const rows = [];
    for (const row of bodyRows) {
        let cells = Array.from(row.querySelectorAll("td"));
        if (cells.length === 0) cells = Array.from(row.querySelectorAll("th"));
        if (cells.length === 0) continue;

        const rowData = {};
        let hasValue = false;
        let pos = 0;
        let cellIdx = 0;
        // Walk column positions left to right until we've placed every
        // real cell this row has AND covered at least the known table
        // width (so trailing active rowSpans still get carried/decremented
        // even if this row has no more of its own cells to place).
        while (cellIdx < cells.length || pos < totalCols) {
            const carry = rowSpanCarry[pos];
            if (carry && carry.rowsLeft > 0) {
                if (pos >= leadingOffset) {
                    const headerIdx = pos - leadingOffset;
                    const header = headerIdx < headers.length ? headers[headerIdx] : ("Column " + (headerIdx + 1));
                    rowData[header] = carry.value;
                    if (carry.value) hasValue = true;
                }
                carry.rowsLeft -= 1;
                pos += 1;
                continue;
            }
            if (cellIdx >= cells.length) {
                pos += 1;
                continue;
            }
            const cell = cells[cellIdx];
            const span = cell.colSpan || 1;
            const rspan = cell.rowSpan || 1;
            const val = cellText(cell);
            if (val) hasValue = true;
            for (let s = 0; s < span; s++) {
                const p = pos + s;
                if (p < leadingOffset) continue;
                const headerIdx = p - leadingOffset;
                const header = headerIdx < headers.length ? headers[headerIdx] : ("Column " + (headerIdx + 1));
                rowData[header] = val;
                if (rspan > 1) {
                    rowSpanCarry[p] = { value: val, rowsLeft: rspan - 1 };
                }
            }
            pos += span;
            cellIdx += 1;
        }
        if (hasValue) rows.push(rowData);
    }

    return { headers: headers, rows: rows, leadingOffset: leadingOffset };
}

function sectionTitleFor(table) {
    // The report page renders each section (Visits / Pharmacies Visits /
    // Office Work / Activities / Events) as a bordered box with a title in
    // the top-left corner -- the classic <fieldset><legend>Title</legend>
    // <table>...</table></fieldset> pattern. If so, this gives us the EXACT
    // section name for a table, which is far more reliable than guessing
    // from row count or header content (e.g. "Visits" vs "Pharmacies Visits"
    // have nearly identical headers and are otherwise hard to tell apart).
    const fieldset = table.closest("fieldset");
    if (fieldset) {
        const legend = fieldset.querySelector("legend");
        if (legend) {
            const text = (legend.innerText || legend.textContent || "").trim();
            if (text) return text;
        }
    }
    // Fallback: look at the nearest preceding heading-like sibling text
    // (in case the real markup isn't a fieldset/legend after all).
    let el = table.previousElementSibling;
    let hops = 0;
    while (el && hops < 5) {
        const text = (el.innerText || el.textContent || "").trim();
        if (text && text.length < 40) return text;
        el = el.previousElementSibling;
        hops++;
    }
    return null;
}

const tables = Array.from(document.querySelectorAll("table"));
return tables.map(t => {
    const extracted = extractTable(t);
    extracted.sectionTitle = sectionTitleFor(t);
    return extracted;
});
"""


def extract_all_tables(driver, debug=False, max_rows_per_table=20000, max_cols_per_table=200):
    print("   📋 Bulk-extracting all tables on the page...")
    t0 = time.time()
    raw_tables = driver.execute_script(EXTRACT_ALL_TABLES_JS)
    print(f"      ✓ Bulk JS extraction finished in {time.time()-t0:.2f}s ({len(raw_tables)} tables found)")

    results = []
    for t_idx, table in enumerate(raw_tables):
        headers = table.get("headers", [])
        rows = table.get("rows", [])
        section_title = table.get("sectionTitle")
        row_count = len(rows)
        col_count = len(headers)
        leading_offset = table.get("leadingOffset", 0)

        if leading_offset > 0:
            print(f"      ⚠️  Table #{t_idx}: detected {leading_offset} unlabeled leading column(s) "
                  f"(e.g. a row-color/index bar with no header) -- auto-corrected the column alignment "
                  f"so real data doesn't shift into the wrong header.")

        # Tables we can positively identify by their fieldset/legend section
        # title (e.g. "Visits", "Office Work") are trusted regardless of row
        # count -- the real Visits table for a multi-user, multi-day pull
        # can easily have 10,000+ rows, and that's NOT junk. The row/col
        # "junk" filter below exists only to catch unrelated giant hidden
        # widget/layout tables that have no identifiable section title.
        has_title = bool(section_title and section_title.strip())
        skipped = (not has_title) and (row_count > max_rows_per_table or col_count > max_cols_per_table)
        if skipped:
            if debug:
                print(f"      ⏭️  Table #{t_idx} skipped ({row_count} rows, {col_count} cols, "
                      f"no section title -- junk widget/layout table)")
            rows = []
        elif debug:
            print(f"      • Table #{t_idx}: title={section_title!r}, {row_count} rows, {col_count} cols")

        results.append({
            "table_index": t_idx,
            "headers": headers,
            "rows": rows,
            "row_count": row_count,
            "leading_offset": leading_offset,
            "col_count": col_count,
            "section_title": section_title,
            "skipped": skipped,
        })

    return results


def find_named_table(tables, expected_titles, debug=False):
    """
    Finds a table whose fieldset/legend section title exactly matches one
    of expected_titles (case/whitespace-insensitive exact match). This is
    the PRIMARY way we identify which physical <table> is "Visits" vs
    "Pharmacies Visits" vs "Office Work" vs "Activities" vs "Events" --
    their header sets are nearly identical in places (e.g. Visits and
    Pharmacies Visits share almost all the same columns), so matching by
    the actual on-page section title beats guessing from headers or row
    count.
    """
    normalized_expected = {normalize_header(t) for t in expected_titles}
    candidates = [
        t for t in tables
        if t.get("section_title") and normalize_header(t["section_title"]) in normalized_expected
    ]
    if not candidates:
        return None
    best = max(candidates, key=lambda t: t["row_count"])
    if debug:
        print(f"      ✓ Found {expected_titles[0]!r} table: #{best['table_index']} "
              f"({best['row_count']} rows, headers: {best['headers']})")
    return best


def find_report_table(tables, debug=False):
    """
    FALLBACK ONLY -- used if find_named_table() couldn't locate a table
    titled "Visits" via fieldset/legend (e.g. the page markup doesn't
    follow that pattern after all). Auto-detects the Visits report table
    by looking for a header row with BOTH a doctor-name-like column AND an
    account-name-like column, while explicitly EXCLUDING tables that look
    like "Pharmacies Visits" (which has near-identical headers plus a few
    pharmacy-only ones) so we don't accidentally grab the wrong table.
    """
    def header_has(headers, keywords):
        joined = [normalize_header(h) for h in headers]
        return any(any(kw in h for kw in keywords) for h in joined)

    PHARMACY_ONLY_KEYWORDS = ["last order", "current stock", "current order"]

    candidates = []
    for t in tables:
        if t["skipped"] or t["row_count"] == 0:
            continue
        headers = t["headers"]
        has_doctor = header_has(headers, COLUMN_KEYWORDS["doctor_name"])
        has_acc = header_has(headers, COLUMN_KEYWORDS["acc_name"])
        is_pharmacy = header_has(headers, PHARMACY_ONLY_KEYWORDS)
        if has_doctor and has_acc and not is_pharmacy:
            candidates.append(t)

    if candidates:
        # If multiple match (shouldn't normally happen), take the one with most rows
        best = max(candidates, key=lambda t: t["row_count"])
        if debug:
            print(f"      ✓ Report table auto-detected: Table #{best['table_index']} "
                  f"({best['row_count']} rows, headers: {best['headers']})")
        return best

    # Last resort: largest surviving table, excluding obvious Pharmacy tables
    surviving = [t for t in tables if not t["skipped"] and t["row_count"] > 0
                 and not header_has(t["headers"], PHARMACY_ONLY_KEYWORDS)]
    if surviving:
        best = max(surviving, key=lambda t: t["row_count"])
        print(f"      ⚠️  Could not confidently identify report table by headers -- "
              f"falling back to largest table (#{best['table_index']}, {best['row_count']} rows). "
              f"Verify this is correct: headers = {best['headers']}")
        return best

    return None


def normalize_header(h):
    """Collapses any whitespace (including newlines from wrapped/multi-line
    header cells like 'Acc.\\nType') into single spaces before matching."""
    return re.sub(r"\s+", " ", (h or "")).strip().lower()


def find_col(headers, keywords):
    """Return the header string that matches one of the keyword synonyms, or None.
    Matching is whitespace-normalized so headers that wrap onto two lines in
    the DOM (e.g. 'Acc.\\nType') still match a keyword like 'acc. type'."""
    for h in headers:
        h_norm = normalize_header(h)
        if any(kw in h_norm for kw in keywords):
            return h
    return None


# ============================================================
# ROW PARSING
# ============================================================

DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")
TIME_RE = re.compile(r"(\d{1,2}:\d{2})")


def parse_date_time(raw_value):
    """Extracts a YYYY-MM-DD date and HH:MM time from a cell that may look
    like '2026-06-10\n09:11' or '2026-06-10 09:11'."""
    date_match = DATE_RE.search(raw_value or "")
    time_match = TIME_RE.search(raw_value or "")
    date_str = date_match.group(1) if date_match else None
    time_str = time_match.group(1) if time_match else None
    return date_str, time_str


def clean_member_name(raw_value):
    """Strips leading '- ' bullet markers etc. from the Members cell.

    Real data format seen in the wild (from platinum_1_one_day_demo.xls):
        "-\\nMohamed Soliman El Sayed"   ← single name with leading dash+newline
        "-\\nEhab Adel"                  ← co-visitor (double visit)
        ""                               ← empty (no co-visitor)

    The cell text is:
        <dash or bullet> [optional space] <name>
    OR
        <dash or bullet> <newline> <name>

    So we must treat a standalone '-' prefix line as a marker to skip, and
    take the FIRST non-empty non-marker line as the name.
    """
    if not raw_value:
        return None
    raw_value = raw_value.replace("\r\n", "\n").replace("\r", "\n").replace("\xa0", " ")
    # Split and strip bullet/dash markers AND standalone "-" lines
    lines = []
    for ln in raw_value.split("\n"):
        stripped = ln.strip(" \t-•·")
        if stripped:        # skip empty and pure-dash lines
            lines.append(stripped)
    return lines[0] if lines else None


def clean_member_list(raw_value):
    """Returns EVERY member name found in the Members cell (not just the first).

    The Members cell in the Visits table contains the NAME OF THE CO-VISITOR
    (the other person present on a double/joint visit), NOT the row's own rep.
    For a single visit the cell is empty or contains just a dash.
    For a double visit it contains '- <manager_name>' or '<rep_name>'.

    This function returns all non-empty, non-marker names found.
    """
    if not raw_value:
        return []
    raw_value = raw_value.replace("\r\n", "\n").replace("\r", "\n").replace("\xa0", " ")
    result = []
    for ln in raw_value.split("\n"):
        stripped = ln.strip(" \t-•·")
        if stripped:
            result.append(stripped)
    return result


def clean_territory(raw_value):
    """Territory cells contain two lines:
       Line 1: territory name  e.g. 'FAYOUM'      ← we want this
       Line 2: city/sub-area   e.g. 'Fayoum City' ← ignore this
    We take only the first non-empty line."""
    if not raw_value:
        return ""
    for line in raw_value.split("\n"):
        line = line.strip()
        if line:
            return line
    return ""


def classify_acc_type(raw_value):
    if not raw_value:
        return "Other"
    key = raw_value.strip().upper()
    return ACC_TYPE_MAP.get(key, "Other")


def classify_visit_type(raw_value):
    """Maps raw Visit Type code (e.g. 'S'/'D') to Single/Double. Unrecognized
    codes map to 'Other' rather than being silently dropped."""
    if not raw_value:
        return "Other"
    key = raw_value.strip().upper()
    return VISIT_TYPE_MAP.get(key, "Other")


def clean_products(raw_value):
    """Products cell is newline-separated with leading '-' markers, e.g.
    '-FLUMOX\\n-VASTAFLAM'. Returns a clean comma-joined string."""
    if not raw_value:
        return ""
    items = [ln.strip(" -\t") for ln in raw_value.split("\n") if ln.strip(" -\t")]
    return ", ".join(items)


def build_records(team_name, table, debug=False):
    """Turns one table's raw rows into a list of clean per-visit records."""
    headers = table["headers"]

    col_member = find_col(headers, COLUMN_KEYWORDS["member"])
    col_acc_type = find_col(headers, COLUMN_KEYWORDS["acc_type"])
    col_acc_name = find_col(headers, COLUMN_KEYWORDS["acc_name"])
    col_acc_id = find_col(headers, COLUMN_KEYWORDS["acc_id"])
    col_doc_id = find_col(headers, COLUMN_KEYWORDS["doc_id"])
    col_doctor_name = find_col(headers, COLUMN_KEYWORDS["doctor_name"])
    col_date = find_col(headers, COLUMN_KEYWORDS["date"])
    col_territory = find_col(headers, COLUMN_KEYWORDS["territory"])
    col_visit_type = find_col(headers, COLUMN_KEYWORDS["visit_type"])
    col_specialty = find_col(headers, COLUMN_KEYWORDS["specialty"])
    col_classification = find_col(headers, COLUMN_KEYWORDS["classification"])
    col_products = find_col(headers, COLUMN_KEYWORDS["products"])
    col_notes = find_col(headers, COLUMN_KEYWORDS["notes"])

    if debug:
        print(f"      Column mapping -> member: {col_member!r}, acc_type: {col_acc_type!r}, "
              f"acc_name: {col_acc_name!r}, acc_id: {col_acc_id!r}, doc_id: {col_doc_id!r}, "
              f"doctor_name: {col_doctor_name!r}, date: {col_date!r}, territory: {col_territory!r}, "
              f"visit_type: {col_visit_type!r}, specialty: {col_specialty!r}, "
              f"classification: {col_classification!r}, products: {col_products!r}")

    records = []
    # --- DIAGNOSTIC: show first 5 raw_member values BEFORE any filtering ---
    # This tells us exactly what the live browser's Members cell contains so
    # we can diagnose why 0 records are parsed even when the table is found.
    if debug:
        print(f"      DIAGNOSTIC - first 5 rows raw field values:")
        for i, row in enumerate(table["rows"][:5]):
            raw_m = row.get(col_member, "") if col_member else "NO_COL_MEMBER"
            raw_d = row.get(col_date, "") if col_date else "NO_COL_DATE"
            raw_a = row.get(col_acc_type, "") if col_acc_type else "NO_COL_ACC_TYPE"
            cleaned = clean_member_name(raw_m)
            print(f"        row[{i}]: Members={repr(raw_m)[:60]}, Date={repr(raw_d)[:30]}, AccType={repr(raw_a)[:20]}, cleaned_user={repr(cleaned)}")

    for row in table["rows"]:
        raw_member = row.get(col_member, "") if col_member else ""
        user = clean_member_name(raw_member)
        if not user:
            continue  # can't attribute this row to anyone
        members = clean_member_list(raw_member)  # all names present on this visit (primary rep + any accompanying manager/coach)

        raw_date = row.get(col_date, "") if col_date else ""
        date_str, time_str = parse_date_time(raw_date)

        raw_acc_type = row.get(col_acc_type, "") if col_acc_type else ""
        acc_type_category = classify_acc_type(raw_acc_type)
        shift = SHIFT_MAP.get(acc_type_category, "Other")  # "AM" / "PM" / "Other"

        raw_visit_type = row.get(col_visit_type, "") if col_visit_type else ""
        visit_type_category = classify_visit_type(raw_visit_type)  # "Single" / "Double" / "Other"

        acc_name = row.get(col_acc_name, "") if col_acc_name else ""
        acc_id = (row.get(col_acc_id, "") if col_acc_id else "").strip()
        doc_id = row.get(col_doc_id, "") if col_doc_id else ""
        doctor_name = row.get(col_doctor_name, "") if col_doctor_name else ""
        territory = clean_territory(row.get(col_territory, "") if col_territory else "")

        # --- Guard against colspan-duplication artifacts ---
        # A merged/spanned cell (e.g. an Office Work or Events row that has
        # no real account, so PulpoPlus collapses several columns into one)
        # now gets its text written into EVERY header slot it visually
        # covers (see EXTRACT_ALL_TABLES_JS) -- that's what stops downstream
        # columns from silently shifting, but it also means Territory can
        # end up holding the SAME echoed text as Acc Type (e.g. both =
        # "Office Work"), which isn't a real territory.
        #
        # NOTE: comparing territory to THIS row's raw_acc_type only (the old
        # check) is fragile -- raw_acc_type can be a short code ("OW") on
        # some pages/rows and the full label ("Office Work") on others, so a
        # literal string-equality check silently fails to fire half the
        # time and "Office Work"/"Events"/etc. leak straight into Territory
        # (confirmed in production output). Instead we check whether the
        # Territory text IS one of the known Acc. Type category labels at
        # all (case-insensitive) -- that catches the artifact regardless of
        # which raw form this particular row happened to use.
        if territory and territory.strip().lower() in ACC_TYPE_LABELS_LOWER:
            territory = ""
        specialty = (row.get(col_specialty, "") if col_specialty else "").strip()
        classification = (row.get(col_classification, "") if col_classification else "").strip()
        products = clean_products(row.get(col_products, "") if col_products else "")
        notes = (row.get(col_notes, "") if col_notes else "").strip()

        # --- Doctor key (deduplication) ---
        # Problem: Two DIFFERENT real doctors can share the same doctor_name +
        # specialty + date because the CRM stores only a generic name (e.g.
        # "Dr Ahmed") while the rep writes the EXACT name in the Notes column.
        # Solution: when Notes is present, fold it into the key so those two
        # visits correctly count as two distinct doctors.
        #   Priority 1 — numeric Doc ID (most reliable)
        #   Priority 2 — name | account (standard fallback)
        #   Priority 3 — name | account | notes (when notes carries the real name)
        doc_id_clean = doc_id.strip()
        doctor_name_clean = doctor_name.strip()
        acc_name_clean = acc_name.strip()
        notes_clean = notes.strip()

        if doc_id_clean:
            doctor_key = doc_id_clean
        elif notes_clean:
            # Notes holds the actual doctor name → use it to distinguish
            doctor_key = f"{doctor_name_clean}|{acc_name_clean}|{notes_clean}"
        else:
            doctor_key = f"{doctor_name_clean}|{acc_name_clean}"

        records.append({
            "team": team_name,
            "user": user,
            "members": members,   # all names present on the visit (for coaching-day detection)
            "territory": territory,
            "date": date_str,
            "time": time_str,
            "acc_type_raw": raw_acc_type,
            "acc_type_category": acc_type_category,
            "shift": shift,
            "visit_type_raw": raw_visit_type,
            "visit_type_category": visit_type_category,
            "acc_id": acc_id,
            "acc_name": acc_name,
            "doctor_key": doctor_key,
            "doctor_name": doctor_name,
            "notes": notes,          # actual doctor name written by rep
            "specialty": specialty,
            "classification": classification,
            "products": products,
        })


    return records


# Column keywords for the SIMPLER Office Work / Activities tables, which
# share an identical schema: # ID Territory Type Shift Date Time Members
# Comments -- no Doctor/Acc Name/Specialty/Products columns exist here at
# all (they genuinely don't apply to office work or generic activities).
SIMPLE_ACTIVITY_COLUMN_KEYWORDS = {
    "member": ["members", "member"],
    "territory": ["territory"],
    "date": ["date"],
    "time": ["time"],
    "comments": ["comments", "comment"],
}


def build_activity_records(team_name, table, category_label, debug=False):
    """
    Parses the Office Work / Activities tables (identical schema:
    # ID Territory Type Shift Date Time Members Comments).

    category_label ("Office Work" or "Activities") comes from the CALLER,
    based on which fieldset/legend section this table was found under --
    NOT re-derived from the row's own "Type" text -- since we already have
    ground truth on which section produced this table.

    shift is always set to "Other" here (even though the table has its own
    native Shift column) so these rows never get counted as an "AM Call" /
    "PM Call" in the doctor-coverage metrics, which are specifically about
    doctor visits. Office Work still correctly counts toward "Office Work
    Days" / fills the AM slot in Total Shifts via acc_type_category, which
    doesn't depend on this shift field.
    """
    headers = table["headers"]

    col_member = find_col(headers, SIMPLE_ACTIVITY_COLUMN_KEYWORDS["member"])
    col_territory = find_col(headers, SIMPLE_ACTIVITY_COLUMN_KEYWORDS["territory"])
    col_date = find_col(headers, SIMPLE_ACTIVITY_COLUMN_KEYWORDS["date"])
    col_time = find_col(headers, SIMPLE_ACTIVITY_COLUMN_KEYWORDS["time"])
    col_comments = find_col(headers, SIMPLE_ACTIVITY_COLUMN_KEYWORDS["comments"])

    if debug:
        print(f"      Column mapping ({category_label}) -> member: {col_member!r}, "
              f"territory: {col_territory!r}, date: {col_date!r}, time: {col_time!r}, "
              f"comments: {col_comments!r}")

    records = []
    for row in table["rows"]:
        raw_member = row.get(col_member, "") if col_member else ""
        user = clean_member_name(raw_member)
        if not user:
            continue
        members = clean_member_list(raw_member)

        territory = clean_territory(row.get(col_territory, "") if col_territory else "")

        # --- Same guard as build_records() ---
        # On this simpler Office Work / Activities schema, some rows render
        # with no real Territory value selected and the CRM echoes the
        # entry's own Type label ("Office Work") into that cell instead of
        # leaving it blank. Left unfixed, that fake "place" was being
        # counted as a genuine extra territory downstream -- which is
        # exactly what was inflating identify_managers()'s per-user
        # territory count and wrongly promoting ordinary reps to "manager"
        # status. Blank it here (same as build_records) so
        # backfill_territory() can recover the user's REAL territory from
        # their other rows afterward, same as the main Visits table.
        if territory and territory.strip().lower() in ACC_TYPE_LABELS_LOWER:
            territory = ""

        raw_date = row.get(col_date, "") if col_date else ""
        raw_time = row.get(col_time, "") if col_time else ""
        date_str, _ = parse_date_time(raw_date)
        _, time_str = parse_date_time(f"{raw_date} {raw_time}")

        notes = (row.get(col_comments, "") if col_comments else "").strip()

        records.append({
            "team": team_name,
            "user": user,
            "members": members,
            "territory": territory,
            "date": date_str,
            "time": time_str,
            "acc_type_raw": category_label,
            "acc_type_category": category_label,
            "shift": "Other",
            "visit_type_raw": None,
            "visit_type_category": "Other",
            "acc_id": None,
            "acc_name": None,
            "doctor_key": None,
            "doctor_name": None,
            "notes": notes,
            "specialty": None,
            "classification": None,
            "products": None,
        })

    return records


# Pharmacies Visits has slightly different columns vs main Visits:
# # Territory Visit-Type Acc.Type Acc.ID Acc.Name Doc.ID Doctor-Name Spec. Class Date Products Last-Order Current-Stock Current-Order Members
# We capture: user (from Members), acc_name (the pharmacy name), acc_id,
# date/time, visit_type, territory, products (full product names), classification.
PHARMACY_COLUMN_KEYWORDS = {
    "member":       ["members", "member"],
    "territory":    ["territory"],
    "acc_name":     ["acc. name", "acc name", "account name"],
    "acc_id":       ["acc. id", "acc id", "account id"],
    "visit_type":   ["visit type"],
    "date":         ["date"],
    "classification": ["class"],
    "products":     ["products", "product"],
}


def build_pharmacy_records(team_name, table, debug=False):
    """
    Parses the 'Pharmacies Visits' table section.
    Each row represents one product-line in a pharmacy visit, so multiple
    rows can share the same pharmacy/date/member (one per product).
    We store every such row separately so:
      - pharmacy_name comes from Acc. Name (NOT Doctor Name)
      - pharmacy_id comes from Acc. ID
      - shift is always "Pharmacy" (separate from AM/PM doctor shifts)
    """
    headers = table["headers"]

    col_member       = find_col(headers, PHARMACY_COLUMN_KEYWORDS["member"])
    col_territory    = find_col(headers, PHARMACY_COLUMN_KEYWORDS["territory"])
    col_acc_name     = find_col(headers, PHARMACY_COLUMN_KEYWORDS["acc_name"])
    col_acc_id       = find_col(headers, PHARMACY_COLUMN_KEYWORDS["acc_id"])
    col_visit_type   = find_col(headers, PHARMACY_COLUMN_KEYWORDS["visit_type"])
    col_date         = find_col(headers, PHARMACY_COLUMN_KEYWORDS["date"])
    col_classification = find_col(headers, PHARMACY_COLUMN_KEYWORDS["classification"])
    col_products     = find_col(headers, PHARMACY_COLUMN_KEYWORDS["products"])

    if debug:
        print(f"      Column mapping (Pharmacies) -> member: {col_member!r}, "
              f"acc_name: {col_acc_name!r}, acc_id: {col_acc_id!r}, "
              f"date: {col_date!r}, products: {col_products!r}")

    records = []
    for row in table["rows"]:
        raw_member = row.get(col_member, "") if col_member else ""
        user = clean_member_name(raw_member)
        if not user:
            continue
        members = clean_member_list(raw_member)

        territory = clean_territory(row.get(col_territory, "") if col_territory else "")
        if territory and territory.strip().lower() in ACC_TYPE_LABELS_LOWER:
            territory = ""

        acc_name  = (row.get(col_acc_name, "") if col_acc_name else "").strip()
        acc_id    = (row.get(col_acc_id, "") if col_acc_id else "").strip()
        raw_date  = row.get(col_date, "") if col_date else ""
        date_str, time_str = parse_date_time(raw_date)

        raw_visit_type = row.get(col_visit_type, "") if col_visit_type else ""
        visit_type_category = classify_visit_type(raw_visit_type)

        classification = (row.get(col_classification, "") if col_classification else "").strip()
        products = clean_products(row.get(col_products, "") if col_products else "")

        records.append({
            "team":               team_name,
            "user":               user,
            "members":            members,
            "territory":          territory,
            "date":               date_str,
            "time":               time_str,
            "acc_type_raw":       "P",
            "acc_type_category":  "Pharmacy",
            "shift":              "Pharmacy",
            "visit_type_raw":     raw_visit_type,
            "visit_type_category": visit_type_category,
            "acc_id":             acc_id,
            "acc_name":           acc_name,    # pharmacy name
            "doctor_key":         None,
            "doctor_name":        None,
            "notes":              None,
            "specialty":          None,
            "classification":     classification,
            "products":           products,
        })

    return records


# ============================================================
# PER-TEAM EXTRACTION LOOP
# ============================================================

def extract_team_records(driver, from_date, to_date, team_id, team_name, debug=False,
                          max_wait=180, settle_time=8, max_rows_per_table=20000, max_cols_per_table=200):
    print(f"\n🔄 Team: {team_name}")

    try:
        driver.get("https://pulpo-eipico.cloud/crm/report_ter_visits_search.php")
        time.sleep(2)

        team_dropdown = driver.find_element(By.NAME, "team")
        select_team = Select(team_dropdown)
        select_team.select_by_value(team_id)
        print("   ✓ Team selected")
        time.sleep(1)

        set_date_field(driver, "dfrom", from_date)
        set_date_field(driver, "dto", to_date)
        print("   ✓ Dates set")
        time.sleep(1)

        # Force checkboxes to exact known state BEFORE every Show click.
        # Only Clinic / Poly Clinics / AM Center / Hospital are checked.
        # Pharmacy, Distributors, Office Work, Activities, Events are unchecked.
        set_acc_type_filters(driver, debug=debug)
        time.sleep(1)

        if not click_show_button(driver):
            return []

        wait_for_full_page_load(driver, max_wait=max_wait, settle_time=settle_time)

        tables = extract_all_tables(driver, debug=debug, max_rows_per_table=max_rows_per_table,
                                     max_cols_per_table=max_cols_per_table)

        # The report page renders FIVE separate tables (Visits / Pharmacies
        # Visits / Office Work / Activities / Events), each with its own
        # header set -- not one unified table with a Type column. We find
        # each one we care about by its fieldset/legend section title.

        # --- Visits: the main doctor-call table (Clinic/Hospital/AM Center/
        # Poly Clinics), with Doctor Name / Acc Name / Specialty / Products ---
        visits_table = find_named_table(tables, ["Visits"], debug=debug)
        if visits_table is None:
            if debug:
                print("      ⚠️  No table titled 'Visits' found via fieldset/legend -- "
                      "falling back to header-based detection")
            visits_table = find_report_table(tables, debug=debug)

        records = []
        if visits_table is not None:
            visit_records = build_records(team_name, visits_table, debug=debug)
            print(f"   ✓ Parsed {len(visit_records)} Visits records (Clinic/Hospital/AM Center/Poly Clinics)")
            records.extend(visit_records)
        else:
            print("   ⚠️  No Visits table found for this team")

        # --- Pharmacies Visits: own table section with pharmacy-specific cols ---
        pharmacy_table = find_named_table(tables, ["Pharmacies Visits", "Pharmacies", "Pharmacy Visits"], debug=debug)
        if pharmacy_table is not None:
            pharmacy_records = build_pharmacy_records(team_name, pharmacy_table, debug=debug)
            print(f"   ✓ Parsed {len(pharmacy_records)} Pharmacy visit records")
            records.extend(pharmacy_records)
        else:
            if debug:
                print("      ℹ️  No 'Pharmacies Visits' table section found")

        # --- Office Work: simpler schema, no doctor/specialty/products ---
        office_table = find_named_table(tables, ["Office Work"], debug=debug)
        if office_table is not None:
            office_records = build_activity_records(team_name, office_table, "Office Work", debug=debug)
            print(f"   ✓ Parsed {len(office_records)} Office Work records")
            records.extend(office_records)

        # --- Activities: identical schema to Office Work ---
        activities_table = find_named_table(tables, ["Activities"], debug=debug)
        if activities_table is not None:
            activity_records = build_activity_records(team_name, activities_table, "Activities", debug=debug)
            print(f"   ✓ Parsed {len(activity_records)} Activities records")
            records.extend(activity_records)

        # --- Events: entirely different schema (no per-row Members/doctor
        # data as such -- attribution is via an "Employee" column instead),
        # and no current KPI depends on it. Not extracted yet, but flagged
        # here rather than silently dropped, so it's a visible known gap.
        events_table = find_named_table(tables, ["Events"], debug=debug)
        if events_table is not None and events_table["row_count"] > 0:
            print(f"   ℹ️  Found an Events table with {events_table['row_count']} rows -- "
                  f"not extracted yet (different schema, no KPI depends on it currently).")

        print(f"   ✓ Parsed {len(records)} total records for this team")
        return records

    except Exception as e:
        print(f"   ❌ Error: {e}")
        return []


def backfill_territory(all_records, debug=False):
    """
    Office Work / Events rows have their Territory cell blanked out by the
    colspan-artifact guard in build_records (it was echoing the Acc. Type
    text, e.g. "Office Work", which isn't a real territory). Leaving it
    blank is still wrong, though -- the rep DOES have a real assigned
    territory that day, it just isn't printed on that particular row. We
    recover it from that SAME user's other rows in this extraction:

        1. Same-date match: if the user has a real (non-blank) territory
           on another row on the SAME calendar date (e.g. a Clinic visit
           earlier that day), use that -- most specific/reliable.
        2. Otherwise: fall back to that user's single most common
           territory across the whole pulled date range.

    Never invents a value out of thin air -- a user with zero known
    territory anywhere in the pull is left blank (extremely rare / would
    indicate a data problem worth investigating separately).
    """
    from collections import Counter

    user_date_territory = {}                      # (user, date) -> first real territory seen
    user_territory_counts = defaultdict(Counter)   # user -> Counter({territory: count})

    for r in all_records:
        if r["territory"]:
            user_territory_counts[r["user"]][r["territory"]] += 1
            key = (r["user"], r["date"])
            if key not in user_date_territory:
                user_date_territory[key] = r["territory"]

    filled_same_date = 0
    filled_majority = 0
    unresolved = 0

    for r in all_records:
        if r["territory"]:
            continue  # already has a real value, nothing to do

        key = (r["user"], r["date"])
        if key in user_date_territory:
            r["territory"] = user_date_territory[key]
            filled_same_date += 1
        elif user_territory_counts[r["user"]]:
            r["territory"] = user_territory_counts[r["user"]].most_common(1)[0][0]
            filled_majority += 1
        else:
            unresolved += 1

    total_filled = filled_same_date + filled_majority
    print(f"   ✓ Territory backfill: {filled_same_date} filled from same-date visit, "
          f"{filled_majority} filled from user's majority territory, "
          f"{unresolved} left blank (no territory known anywhere for that user)")
    if debug and unresolved:
        unresolved_users = sorted(set(r["user"] for r in all_records if not r["territory"]))
        print(f"      ⚠️  Users with NO territory anywhere in the pull: {unresolved_users}")

    return total_filled


# ============================================================
# PER-USER KPI SUMMARY
# ============================================================

def minutes_between(t1, t2):
    """t1, t2 are 'HH:MM' strings. Returns absolute difference in minutes."""
    fmt = "%H:%M"
    d1 = datetime.strptime(t1, fmt)
    d2 = datetime.strptime(t2, fmt)
    return abs((d2 - d1).total_seconds()) / 60.0


def identify_managers(all_records):
    """
    A user counts as a manager/coach if their visit records span MORE THAN
    ONE distinct territory (normal reps are assigned a single territory).
    e.g. "Mohamad Elhady    CAIRO 1; CAIRO 2; CAIRO 3; CAIRO 4" -> manager.
    This is derived straight from the extracted data -- no separate lookup
    table is required.

    Defense-in-depth: territory should already be a REAL place by the time
    this runs (backfill_territory() is called first in main() and the
    build_records/build_activity_records guards blank out "Office Work" /
    "Events" / etc. before that). But we additionally filter out anything
    that still matches a known Acc. Type category label here too, so a
    single unresolved/edge-case row can never masquerade as an extra
    "place" and wrongly promote an ordinary single-territory rep to
    manager status.
    """
    territories_by_user = defaultdict(set)
    for r in all_records:
        t = r["territory"]
        if t and t.strip().lower() not in ACC_TYPE_LABELS_LOWER:
            territories_by_user[r["user"]].add(t)
    return {u for u, terrs in territories_by_user.items() if len(terrs) > 1}


COACHING_TIME_TOLERANCE_MINUTES = 30  # how close two visit times must be to count as "the same visit"


def compute_coaching_days(all_records, managers, debug=False):
    """
    Coaching day = a date where a manager did THE SAME visits as a rep for
    BOTH the AM shift AND the PM shift that day.

    In practice the manager and rep are each logged in as their own user,
    so a joint visit shows up as TWO SEPARATE visit records -- one under
    the manager's name, one under the rep's -- to the SAME account
    (acc_id), on the same date and shift, at almost the same time (not one
    shared "Members" cell listing both people). So detection is:

      1. Manager and rep both have a visit record for the SAME acc_id
      2. On the SAME date and SAME shift (AM or PM)
      3. Within COACHING_TIME_TOLERANCE_MINUTES of each other

    If that holds for BOTH the AM shift and the PM shift on the same date,
    that date counts as ONE coaching day for that manager (with that rep).

    (We also still honor the older signal -- a visit row whose Members
    cell explicitly lists both the manager and the rep together -- in case
    any rows are entered that way; either signal can satisfy a shift.)

    Returns:
      coaching_days_by_manager : {manager_name: coaching_day_count}
      detail_df                : one row per (Manager, Rep, Date) coaching day
    """
    # (manager, rep, date) -> set of shifts ("AM"/"PM") confirmed as coached
    matched_shifts = defaultdict(set)

    # --- Signal 1: separate manager/rep records hitting the same account,
    # same date+shift, at almost the same time ---
    # acc_id is normalized (.strip()) both here and at parse time --
    # a stray trailing space/newline difference between the manager's and
    # rep's copy of the SAME account cell was enough to silently break this
    # dict key match and drop a real matched shift, undercounting the day.
    visits_by_key = defaultdict(list)  # (date, shift, acc_id) -> [(user, time_str), ...]
    for r in all_records:
        if r["shift"] not in ("AM", "PM"):
            continue
        acc_id_norm = (r["acc_id"] or "").strip()
        if not r["date"] or not acc_id_norm:
            continue
        visits_by_key[(r["date"], r["shift"], acc_id_norm)].append((r["user"], r["time"]))

    for (date, shift, acc_id), visits in visits_by_key.items():
        manager_visits = [(u, t) for u, t in visits if u in managers]
        if not manager_visits:
            continue
        rep_visits = [(u, t) for u, t in visits if u not in managers]
        for m_user, m_time in manager_visits:
            for rep_user, rep_time in rep_visits:
                if rep_user == m_user:
                    continue
                if m_time and rep_time:
                    if minutes_between(m_time, rep_time) > COACHING_TIME_TOLERANCE_MINUTES:
                        continue  # same account, but too far apart in time -- probably coincidence, not a joint visit
                matched_shifts[(m_user, rep_user, date)].add(shift)

    # --- Signal 2 (fallback): a visit row whose Members cell lists both
    # the manager and the rep together (single combined row, not two
    # separate ones) ---
    # BUG FIXED: this used to assume r["user"] (the row's primary name,
    # i.e. whichever name PulpoPlus happens to print FIRST in the Members
    # cell) is always the REP, then searched for a manager among the
    # remaining names. That silently breaks whenever the MANAGER happens
    # to be the one listed first -- r["user"] becomes the manager, so the
    # code was scanning the rest of the list for ANOTHER manager, found
    # none, and dropped a real coaching visit entirely (confirmed: this is
    # exactly what happened for Abdelkarem Gadelrab on 2026-06-04, where
    # every row read Members = "Abdelkarem Gadelrab, Mariana Moner Morad"
    # -- manager first -- and every one of that day's AM+PM shifts was
    # silently dropped). Fixed by pairing every manager present with
    # every non-manager present in the row, regardless of ordering / which
    # one is the row's primary user.
    for r in all_records:
        if r["shift"] not in ("AM", "PM") or not r["date"]:
            continue
        members = r.get("members") or [r["user"]]
        if len(members) < 2:
            continue
        managers_present = [m for m in members if m in managers]
        reps_present = [m for m in members if m not in managers]
        if not managers_present or not reps_present:
            continue
        for m in managers_present:
            for rep in reps_present:
                matched_shifts[(m, rep, r["date"])].add(r["shift"])

    coaching_days_by_manager = defaultdict(int)
    detail_rows = []
    partial_matches = []  # (manager, rep, date, shifts) where only ONE of AM/PM matched
    for (manager, rep, date), shifts in matched_shifts.items():
        if {"AM", "PM"}.issubset(shifts):
            coaching_days_by_manager[manager] += 1
            detail_rows.append({"Manager": manager, "Rep": rep, "Date": date})
        else:
            partial_matches.append((manager, rep, date, shifts))

    if debug and partial_matches:
        print(f"\n   🔍 COACHING DAYS -- partial matches (only ONE shift confirmed, "
              f"so these dates did NOT count as a full coaching day):")
        for manager, rep, date, shifts in sorted(partial_matches):
            print(f"      {manager} + {rep} on {date}: only {sorted(shifts)} matched "
                  f"(needs both AM and PM to count)")
        print(f"      -> If you expected one of these to be a full coaching day, check "
              f"that BOTH the manager's and the rep's row for that missing shift share "
              f"the exact same Acc. ID on that date -- that's the only thing that can "
              f"stop a real joint visit from being detected here.\n")

    detail_rows.sort(key=lambda x: (x["Manager"], x["Date"], x["Rep"]))
    detail_df = pd.DataFrame(detail_rows)
    return dict(coaching_days_by_manager), detail_df


def compute_summary(all_records, debug=False, managers=None, coaching_days_by_manager=None):
    if debug:
        raw_acc_types_seen = sorted(set(r["acc_type_raw"] for r in all_records if r["acc_type_raw"]))
        print(f"\n🔍 RAW ACC TYPE VALUES SEEN ON PAGE: {raw_acc_types_seen}")
        print("   (Compare against ACC_TYPE_MAP at the top of the script -- "
              "fix/extend it if any of these aren't mapping to the category you expect)")

        raw_visit_types_seen = sorted(set(r["visit_type_raw"] for r in all_records if r["visit_type_raw"]))
        print(f"\n🔍 RAW VISIT TYPE VALUES SEEN ON PAGE: {raw_visit_types_seen}")
        print("   (Compare against VISIT_TYPE_MAP at the top of the script -- "
              "fix/extend it if 'Double' visits aren't being recognized)\n")

    by_user = defaultdict(list)
    for r in all_records:
        by_user[r["user"]].append(r)

    summary_rows = []

    for user, records in sorted(by_user.items()):
        # --- Territory ---
        # Defense-in-depth: exclude any lingering Acc. Type category label
        # (e.g. "Office Work") the same way identify_managers() does, so
        # the displayed Territory column can never show a fake "place"
        # even in an edge case backfill_territory() couldn't resolve.
        territories = sorted(set(
            r["territory"] for r in records
            if r["territory"] and r["territory"].strip().lower() not in ACC_TYPE_LABELS_LOWER
        ))
        territory_str = "; ".join(territories)

        # --- Manager identification ---
        # Normal reps are assigned to a single territory. Managers/coaches
        # cover multiple territories (e.g. "CAIRO 1; CAIRO 2; CAIRO 3;
        # CAIRO 4"), so we use ">1 distinct territory" as the signal that
        # this user is a manager, rather than needing a separate lookup table.
        # Uses the same `managers` set computed once (by identify_managers)
        # and shared with compute_coaching_days, for consistency; falls back
        # to a local territory-count check if compute_summary is called
        # standalone without that set.
        is_manager = (user in managers) if managers is not None else (len(territories) > 1)

        # --- Working days ---
        # A "working day" = a calendar date where the rep had at least 1 visit.
        dates_with_visits = sorted(set(r["date"] for r in records if r["date"]))
        working_days = len(dates_with_visits)

        # --- Shift count & effective working days ---
        # Each full working day ideally has 1 AM shift + 1 PM shift.
        # On some days the rep may only have a PM shift (office morning) or
        # only an AM shift. We count every date on which the rep had ≥1
        # AM-shift visit as 1 AM shift day, same for PM. Dividing the total
        # by 2 gives the number of "complete field days" (both shifts worked).
        #   Total Shifts = AM shift days + PM shift days
        #   Complete Field Days = Total Shifts ÷ 2
        am_shift_dates = set(r["date"] for r in records if r["date"] and r["shift"] == "AM")
        pm_shift_dates = set(r["date"] for r in records if r["date"] and r["shift"] == "PM")
        am_shift_days = len(am_shift_dates)   # pure field AM (Hospital + AM Center) days
        pm_shift_days = len(pm_shift_dates)

        # Office Work can stand IN FOR the AM shift on a given day (e.g. the
        # rep does office/admin work in the morning, then a normal PM field
        # shift). Those dates still count as a filled "AM slot" for the
        # purposes of Total Shifts / Complete Field Days, even though they
        # contribute 0 to AM doctor-coverage metrics.
        office_work_dates = set(
            r["date"] for r in records
            if r["date"] and r["acc_type_category"] == "Office Work"
        )
        office_work_days = len(office_work_dates)
        am_slot_dates = am_shift_dates | office_work_dates
        am_slot_days = len(am_slot_dates)   # AM field shift OR Office Work, whichever filled that morning

        total_shifts = am_slot_days + pm_shift_days
        complete_field_days = round(total_shifts / 2, 1)   # e.g. 9 shifts → 4.5

        # --- Double-visit days (kept for reference) ---
        # A date where at least 1 visit row has visit_type == "Double".
        double_visit_dates = set(
            r["date"] for r in records
            if r["date"] and r["visit_type_category"] == "Double"
        )
        double_visit_days = len(double_visit_dates)

        # --- Coaching days (managers only) ---
        # A coaching day = a date where the manager accompanied the SAME rep
        # for BOTH the AM shift AND the PM shift on that date.
        # coaching_days_by_manager is precomputed once (across ALL users'
        # records together, since it requires comparing two users' visits)
        # by compute_coaching_days() and passed in here.
        coaching_days = (coaching_days_by_manager or {}).get(user, 0)

        # --- Covered doctors per account type (unique doctor_key per category) ---
        doctors_by_category = defaultdict(set)
        for r in records:
            if r["acc_type_category"] in COVERAGE_ACC_TYPES and r["doctor_key"]:
                doctors_by_category[r["acc_type_category"]].add(r["doctor_key"])

        total_covered_doctors = sum(len(v) for v in doctors_by_category.values())

        # --- AM / PM splits (AM = Hospital + AM Center, PM = Clinic + Poly Clinics) ---
        am_records = [r for r in records if r["shift"] == "AM"]
        pm_records = [r for r in records if r["shift"] == "PM"]

        am_unique_doctors = set(r["doctor_key"] for r in am_records if r["doctor_key"])
        pm_unique_doctors = set(r["doctor_key"] for r in pm_records if r["doctor_key"])

        am_specialties = set(r["specialty"] for r in am_records if r["specialty"])
        pm_specialties = set(r["specialty"] for r in pm_records if r["specialty"])

        am_calls = len(am_records)   # total calls/visits, not unique doctors
        pm_calls = len(pm_records)

        # --- Events / Office Work counts ---
        no_events = sum(1 for r in records if r["acc_type_category"] == "Events")
        no_office_work = sum(1 for r in records if r["acc_type_category"] == "Office Work")

        # --- Product call counts ---
        # Each visit row may list one or more products (comma-joined after cleaning).
        # We count how many visits (calls) mentioned each product, across AM+PM
        # doctor visits only (shift != "Other").
        product_call_counts = defaultdict(int)
        for r in records:
            if r["shift"] == "Other":
                continue  # skip events / office work / pharmacy rows
            if not r["products"]:
                continue
            for prod in r["products"].split(","):
                prod = prod.strip()
                if prod:
                    product_call_counts[prod] += 1
        # Sorted descending by call count so most-promoted product comes first
        products_summary = ", ".join(
            f"{prod}({cnt})" for prod, cnt in
            sorted(product_call_counts.items(), key=lambda x: -x[1])
        )
        total_product_calls = sum(product_call_counts.values())
        distinct_products = len(product_call_counts)

        # --- AM / PM shift duration per day ---
        # AM shift: first visit of the day → last visit of the day among AM records.
        # PM shift: same logic for PM records.
        # We average this duration across all days the user had that shift.
        def avg_shift_duration(shift_records):
            by_day_times = defaultdict(list)
            for r in shift_records:
                if r["date"] and r["time"]:
                    by_day_times[r["date"]].append(r["time"])
            daily_durations = []
            for date, times in by_day_times.items():
                times_sorted = sorted(times)
                if len(times_sorted) >= 2:
                    daily_durations.append(minutes_between(times_sorted[0], times_sorted[-1]))
                else:
                    daily_durations.append(0.0)
            return (sum(daily_durations) / len(daily_durations)) if daily_durations else 0.0

        def fmt_hm(total_minutes):
            h = int(total_minutes // 60)
            m = int(total_minutes % 60)
            return f"{h}:{m:02d}"

        def avg_first_visit_clock_time(shift_records):
            """Average clock time (e.g. '08:45') of the FIRST visit of the
            day, across every day the user had that shift. This answers
            "what time does this rep typically start their AM shift?" --
            different from avg_shift_duration, which measures how LONG the
            shift lasted, not when it started."""
            by_day_times = defaultdict(list)
            for r in shift_records:
                if r["date"] and r["time"]:
                    by_day_times[r["date"]].append(r["time"])
            first_visit_minutes = []
            for date, times in by_day_times.items():
                first_time = sorted(times)[0]
                try:
                    dt = datetime.strptime(first_time, "%H:%M")
                    first_visit_minutes.append(dt.hour * 60 + dt.minute)
                except Exception:
                    continue
            if not first_visit_minutes:
                return None
            avg_minutes = round(sum(first_visit_minutes) / len(first_visit_minutes))
            h, m = divmod(avg_minutes, 60)
            return f"{h:02d}:{m:02d}"

        am_avg_duration_min = avg_shift_duration(am_records)
        pm_avg_duration_min = avg_shift_duration(pm_records)
        am_avg_start_time = avg_first_visit_clock_time(am_records)   # e.g. "08:45"

        # --- Overall avg field duration (all visit types combined) ---
        overall_avg_min = avg_shift_duration(records)

        # --- Pharmacy visits (Req 2) ---
        # Each row in pharmacy_records is one product-line visit to a pharmacy.
        # "Pharmacies Visited" = total pharmacy visit ROWS (calls/lines).
        # "Pharmacies Covered" = count of UNIQUE pharmacy accounts (by acc_id or name).
        pharmacy_records = [r for r in records if r["acc_type_category"] == "Pharmacy"]
        pharmacies_visited = len(pharmacy_records)  # total rows = total pharmacy visits
        pharmacy_unique_ids = set()
        for r in pharmacy_records:
            key = r["acc_id"] if r["acc_id"] else r["acc_name"]
            if key:
                pharmacy_unique_ids.add(key)
        pharmacies_covered = len(pharmacy_unique_ids)  # unique pharmacy accounts reached

        # --- PM Call Rate (total visits ÷ PM working days) ---
        # Definition: TOTAL PM visits (all rows, including revisits to same
        # doctor) divided by the number of days the user had a PM shift.
        # Example: 10 unique doctors but 12 visits (2 revisited) over 5 PM
        # days → PM Call Rate = 12 ÷ 5 = 2.4
        # This is pm_calls (count of all PM rows) ÷ pm_shift_days.
        pm_call_rate = round(pm_calls / pm_shift_days, 1) if pm_shift_days > 0 else 0.0
        am_call_rate = round(am_calls / am_shift_days, 1) if am_shift_days > 0 else 0.0

        summary_rows.append({
            # ── Identity ──────────────────────────────────────────────────────
            "User": user,
            "Territory": territory_str,

            # ── Activity / Field Days ─────────────────────────────────────────
            "Working Days": working_days,
            "Complete Field Days": complete_field_days,
            "AM Shift Days": am_shift_days,
            "PM Shift Days": pm_shift_days,
            "Double Visit Days": double_visit_days,
            "Coaching Days": coaching_days,
            "Office Work Days": office_work_days,

            # ── Doctor coverage breakdown ──────────────────────────────────────
            "Clinic Doctors Covered": len(doctors_by_category.get("Clinic", set())),
            "Poly Clinic Doctors Covered": len(doctors_by_category.get("Poly Clinics", set())),
            "Total PM Covered": (len(doctors_by_category.get("Clinic", set()))
                                  + len(doctors_by_category.get("Poly Clinics", set()))),
            "AM Center Doctors Covered": len(doctors_by_category.get("AM Center", set())),
            "Hospital Doctors Covered": len(doctors_by_category.get("Hospital", set())),
            "Total AM Covered": (len(doctors_by_category.get("AM Center", set()))
                                  + len(doctors_by_category.get("Hospital", set()))),

            # ── Calls & Call Rate ────────────────────────────────────────────
            "PM Calls": pm_calls,              # total PM visits incl. revisits
            "PM Shift Days": pm_shift_days,
            "PM Call Rate": pm_call_rate,      # total PM visits ÷ PM working days
            "AM Calls": am_calls,
            "AM Shift Days": am_shift_days,
            "AM Call Rate": am_call_rate,      # total AM visits ÷ AM working days
            "Avg AM Starting Time": am_avg_start_time,

            # ── Pharmacy (Req 2) ─────────────────────────────────────────────
            "Pharmacies Visited": pharmacies_visited,   # total pharmacy call rows
            "Pharmacies Covered": pharmacies_covered,   # unique pharmacy accounts

            # ── Shift durations ─────────────────────────────────────────────
            "Avg AM Shift Duration (h:mm)": fmt_hm(am_avg_duration_min),
            "Avg PM Shift Duration (h:mm)": fmt_hm(pm_avg_duration_min),
            "Avg Field Duration Overall (h:mm)": fmt_hm(overall_avg_min),
        })

    return pd.DataFrame(summary_rows)


def compute_specialty_classification(all_records):
    """
    Tidy breakdown: one row per (User, Specialty, Classification, Shift)
    combination, with visit counts and unique-doctor counts. This lets you
    answer questions like "how many Class A gynecology calls did this user
    make in the AM shift" by filtering/pivoting this sheet in Excel.
    """
    groups = defaultdict(lambda: {"visits": 0, "doctors": set()})

    for r in all_records:
        if not r["specialty"]:
            continue  # skip rows with no specialty (events, office work, etc.)
        key = (r["user"], r["specialty"], r["classification"] or "(Unclassified)", r["shift"])
        groups[key]["visits"] += 1
        if r["doctor_key"]:
            groups[key]["doctors"].add(r["doctor_key"])

    rows = []
    for (user, specialty, classification, shift), agg in sorted(
        groups.items(), key=lambda kv: tuple(str(x) for x in kv[0])
    ):
        rows.append({
            "User": user,
            "Specialty": specialty,
            "Classification": classification,
            "Shift": shift,
            "Visit Count": agg["visits"],
            "Unique Doctors": len(agg["doctors"]),
        })

    return pd.DataFrame(rows)


# ============================================================
# SAVE
# ============================================================

def compute_product_calls(all_records):
    """
    One row per (User, Product, Shift, Specialty) -- i.e. Specialty is
    UNPIVOTED into its own column (long/tidy format) rather than bundled
    into one comma-joined cell, so this sheet can be dropped straight into
    a BI tool (Power BI / Tableau) and used as a dimension for bar/pie
    charts without any extra transform step.

      - Call Count    : number of visits where that product was promoted to
                        that specialty
      - Unique Doctors: distinct doctors of that specialty reached with
                        that product

    This lets you answer: "User X made 12 calls on FLUMOX to Gynecology
    doctors in the PM shift, and 8 to Andrology."
    Only AM/PM (doctor) visits are included; Events/OfficeWork/Other are
    skipped. Visits with no specialty recorded are grouped under
    "(Unspecified)" rather than dropped, so totals still reconcile.
    """
    groups = defaultdict(lambda: {"calls": 0, "doctors": set()})

    for r in all_records:
        if r["shift"] == "Other":
            continue
        if not r["products"]:
            continue
        specialty = r["specialty"].strip() if r["specialty"] else "(Unspecified)"
        for prod in r["products"].split(","):
            prod = prod.strip()
            if not prod:
                continue
            key = (r["user"], prod, r["shift"], specialty)
            groups[key]["calls"] += 1
            if r["doctor_key"]:
                groups[key]["doctors"].add(r["doctor_key"])

    rows = []
    for (user, product, shift, specialty), agg in sorted(
        groups.items(), key=lambda kv: tuple(str(x) for x in kv[0])
    ):
        rows.append({
            "User": user,
            "Product": product,
            "Shift": shift,
            "Specialty": specialty,
            "Call Count": agg["calls"],
            "Unique Doctors": len(agg["doctors"]),
        })

    return pd.DataFrame(rows)


def style_worksheet(ws, num_data_rows, num_cols):
    """
    Applies consistent formatting to every sheet so rows are easy to tell
    apart at a glance:
      - Bold, colored header row (frozen so it stays visible while scrolling)
      - Alternating row banding (light fill on every other data row)
      - Thin borders around every cell
      - Auto-sized column widths (sampled, so this stays fast on huge sheets)
    """
    if num_cols == 0 or num_data_rows == 0:
        return

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    band_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    thin = Side(style="thin", color="B7C6D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Header row ---
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # --- Banded data rows (alternate fill every other row for differentiation) ---
    for row in range(2, num_data_rows + 2):
        banded = (row % 2 == 0)
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            if banded:
                cell.fill = band_fill
            cell.border = border

    # --- Freeze header row so it stays visible while scrolling ---
    ws.freeze_panes = "A2"

    # --- Auto-size columns (sampled for speed on very large sheets) ---
    sample_rows = min(num_data_rows + 1, 200)
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, sample_rows + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def save_results(summary_df, all_records, from_date, to_date, team_tag=None,
                 specialty_class_df=None, product_df=None, coaching_detail_df=None):
    date_str = f"{from_date.replace('-', '')}_{to_date.replace('-', '')}"
    suffix = f"_{team_tag}" if team_tag else ""
    filename = f"pulpoplus_user_summary_{date_str}{suffix}.xlsx"

    print(f"\n💾 Saving to {filename}...")

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # Sheet 1 — Main KPI summary (one row per user)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        style_worksheet(writer.sheets["Summary"], len(summary_df), len(summary_df.columns))
        print(f"   ✓ Summary sheet: {len(summary_df)} users")

        # Sheet 2 — Specialty × Classification breakdown (one row per user/specialty/class/shift)
        if specialty_class_df is not None and not specialty_class_df.empty:
            specialty_class_df.to_excel(writer, sheet_name="Specialty x Class", index=False)
            style_worksheet(writer.sheets["Specialty x Class"], len(specialty_class_df), len(specialty_class_df.columns))
            print(f"   ✓ Specialty x Class sheet: {len(specialty_class_df)} rows")

        # Sheet 3 — Product calls, unpivoted per specialty (BI-ready long format)
        if product_df is not None and not product_df.empty:
            product_df.to_excel(writer, sheet_name="Product Calls per spec", index=False)
            style_worksheet(writer.sheets["Product Calls per spec"], len(product_df), len(product_df.columns))
            print(f"   ✓ Product Calls per spec sheet: {len(product_df)} rows")

        # Sheet 4 — Coaching days detail (one row per manager/rep/date where
        # the manager stayed with the same rep for the whole AM+PM day)
        if coaching_detail_df is not None and not coaching_detail_df.empty:
            coaching_detail_df.to_excel(writer, sheet_name="Coaching Days", index=False)
            style_worksheet(writer.sheets["Coaching Days"], len(coaching_detail_df), len(coaching_detail_df.columns))
            print(f"   ✓ Coaching Days sheet: {len(coaching_detail_df)} rows")

        # Sheet 5 — Full raw visit log
        raw_df = pd.DataFrame(all_records)
        if not raw_df.empty:
            # Req 3: In Raw Data the "members" column shows only the ONE
            # name that is the row's own user -- not the full list of
            # everyone on a joint visit. The full list is used internally
            # for coaching-day detection (above) but there's no need to
            # expose it here, and seeing "Abel Allaha, Abdelkarem Gadelrab"
            # on every double-visit row is confusing when you're looking at
            # Abd Allaha's report. The user is already stored as "user",
            # so we simply echo that.
            raw_df["members"] = raw_df["user"]

            if len(raw_df) > 1_000_000:
                raw_df = raw_df.iloc[:1_000_000]
            raw_df.to_excel(writer, sheet_name="Raw Data", index=False)
            style_worksheet(writer.sheets["Raw Data"], len(raw_df), len(raw_df.columns))
            print(f"   ✓ Raw Data sheet: {len(raw_df)} visit records")

    print(f"\n✅ File saved: {filename}")
    return filename


# ============================================================
# MAIN
# ============================================================

def main():
    print("\n" + "=" * 70)
    print("🔍 PulpoPlus Extractor - SMART SUMMARY VERSION")
    print("=" * 70 + "\n")

    args = parse_arguments()
    from_date, to_date = get_date_range(args.from_date, args.to_date)

    print(f"📅 Date: {from_date} to {to_date}")
    print(f"🐛 Debug mode: {args.debug}")
    print(f"⏳ Max wait: {args.max_wait}s, settle time: {args.settle_time}s\n")

    driver = setup_driver()
    all_records = []

    try:
        if not login(driver):
            return

        teams = navigate_to_reports(driver)
        if not teams:
            return

        if args.list_teams:
            print("\n📋 Available teams:")
            for t in teams:
                print(f"   - {t['name']}")
            return

        if args.team:
            wanted = args.team.strip().lower()
            matched = [t for t in teams if wanted in t['name'].strip().lower()]
            if not matched:
                print(f"\n❌ No team matched '{args.team}'. Available teams:")
                for t in teams:
                    print(f"   - {t['name']}")
                return
            if len(matched) > 1:
                print(f"\n⚠️  '{args.team}' matched multiple teams, using the first match. "
                      f"All matches: {[t['name'] for t in matched]}")
            teams = matched[:1]
            print(f"\n🎯 TEST MODE: running on single team only -> {teams[0]['name']}")

        for idx, team in enumerate(teams, 1):
            print(f"\n[{idx}/{len(teams)}] {team['name']}")
            records = extract_team_records(
                driver, from_date, to_date, team['value'], team['name'],
                debug=args.debug, max_wait=args.max_wait, settle_time=args.settle_time,
                max_rows_per_table=args.max_rows_per_table, max_cols_per_table=args.max_cols_per_table
            )
            all_records.extend(records)

        if not all_records:
            print("\n❌ No visit records were extracted -- nothing to summarize")
            return

        print(f"\n📊 Computing per-user summary from {len(all_records)} total visit records...")

        # Recover the real territory for Office Work / Events rows whose
        # Territory cell was blanked by the colspan-artifact guard.
        backfill_territory(all_records, debug=args.debug)

        managers = identify_managers(all_records)
        if args.debug:
            print(f"   👤 Managers/coaches identified (multi-territory users): {sorted(managers) or '(none)'}")
        coaching_days_by_manager, coaching_detail_df = compute_coaching_days(all_records, managers, debug=args.debug)

        summary_df          = compute_summary(all_records, debug=args.debug, managers=managers,
                                               coaching_days_by_manager=coaching_days_by_manager)
        specialty_class_df  = compute_specialty_classification(all_records)
        product_df          = compute_product_calls(all_records)           # "Product Calls per spec" (unpivoted, BI-ready)

        team_tag = None
        if args.team:
            safe_name = re.sub(r"[^A-Za-z0-9]+", "", teams[0]['name'])
            team_tag = f"TEST_{safe_name}"

        save_results(summary_df, all_records, from_date, to_date, team_tag=team_tag,
                     specialty_class_df=specialty_class_df, product_df=product_df,
                     coaching_detail_df=coaching_detail_df)

    finally:
        driver.quit()
        print("\n🌐 Browser closed")


if __name__ == "__main__":
    main()