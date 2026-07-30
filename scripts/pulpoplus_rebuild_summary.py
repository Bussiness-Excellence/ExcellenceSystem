#!/usr/bin/env python3
"""
pulpoplus_rebuild_summary.py
────────────────────────────
Reads the "Raw Data" sheet from an existing pulpoplus_user_summary_*.xlsx
file and regenerates every other sheet (Summary, Specialty x Class,
Product Calls per spec, Coaching Days) exactly as the extractor would.

Use this when:
  • You already ran the extractor and have the Excel file
  • You want to recalculate summaries after a logic fix
  • You want to combine Raw Data from several extractions into one summary
  • You don't want to re-run the browser automation

Usage:
    python pulpoplus_rebuild_summary.py input.xlsx
    python pulpoplus_rebuild_summary.py input.xlsx --out rebuilt.xlsx
    python pulpoplus_rebuild_summary.py raw1.xlsx raw2.xlsx raw3.xlsx   # merge multiple files
    python pulpoplus_rebuild_summary.py input.xlsx --debug
"""

import sys
import re
import argparse
import os
from collections import defaultdict, Counter
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── optional hierarchy integration ───────────────────────────────────────────
try:
    import pulpoplus_hierarchy as _hier
    _HIERARCHY_AVAILABLE = True
except ImportError:
    _HIERARCHY_AVAILABLE = False

def _load_hierarchy_if_available(path=None):
    """
    Load the 2026 team-structure workbook if available.

    Resolution order:
      1. Explicit --hierarchy path, if given.
      2. Any .xlsx file inside a "hierarchy" folder next to this script
         (or in the current directory) — filename doesn't matter, so you
         can just drop a replacement file in there (even with a different
         name, e.g. after re-downloading) and it's picked up automatically.
         If more than one file is in there, the most recently modified one
         wins (a note is printed either way).
      3. Falls back to the old exact-filename check in the current/script
         directory, for backward compatibility.

    Returns a HierarchyMap or None.
    """
    if not _HIERARCHY_AVAILABLE:
        return None
    if path and os.path.exists(path):
        return _hier.load_hierarchy(path)

    script_dir = os.path.dirname(os.path.abspath(__file__))

    # 2. Dedicated "hierarchy" folder — any .xlsx file in it, regardless of name
    for search_dir in (os.getcwd(), script_dir):
        folder = os.path.join(search_dir, "hierarchy")
        if os.path.isdir(folder):
            xlsx_files = [
                os.path.join(folder, f) for f in os.listdir(folder)
                if f.lower().endswith(".xlsx") and not f.startswith("~$")
            ]
            if len(xlsx_files) == 1:
                print(f"  📋 Auto-discovered hierarchy file: {xlsx_files[0]}")
                return _hier.load_hierarchy(xlsx_files[0])
            elif len(xlsx_files) > 1:
                newest = max(xlsx_files, key=os.path.getmtime)
                print(f"  📋 Found {len(xlsx_files)} files in '{folder}' — "
                      f"using the most recently modified: {newest}")
                print(f"     (tip: keep only one file in that folder to avoid ambiguity)")
                return _hier.load_hierarchy(newest)

    # 3. Old exact-filename fallback in cwd/script_dir
    candidates = [
        "team_structure_2026.xlsx",
        "NEW_2026_Promotional_Team_Structure.xlsx",
    ]
    for candidate in candidates:
        for search_dir in (os.getcwd(), script_dir):
            full = os.path.join(search_dir, candidate)
            if os.path.exists(full):
                print(f"  📋 Auto-discovered hierarchy file: {full}")
                return _hier.load_hierarchy(full)
    return None

# ── same logic constants as the extractor ────────────────────────────────────

ACC_TYPE_MAP = {
    "C": "Clinic", "CLINIC": "Clinic",
    "P": "Pharmacy", "PH": "Pharmacy", "PHARMACY": "Pharmacy",
    "H": "Hospital", "HOSPITAL": "Hospital",
    "AM": "AM Center", "AM CENTER": "AM Center",
    "PC": "Poly Clinics", "POLY CLINIC": "Poly Clinics", "POLY CLINICS": "Poly Clinics",
    "D": "Distributors", "DISTRIBUTOR": "Distributors", "DISTRIBUTORS": "Distributors",
    "OW": "Office Work", "OFFICE WORK": "Office Work",
    "A": "Activities", "ACTIVITY": "Activities", "ACTIVITIES": "Activities",
    "E": "Events", "EVENT": "Events", "EVENTS": "Events",
}
ACC_TYPE_LABELS_LOWER = {k.lower() for k in ACC_TYPE_MAP} | {v.lower() for v in ACC_TYPE_MAP.values()}

SHIFT_MAP = {
    "Clinic":       "PM",
    "Poly Clinics": "PM",
    "Hospital":     "AM",
    "AM Center":    "AM",
}
COACHING_TIME_TOLERANCE_MINUTES = 90

# What actually counts as a "call": AM shift = Hospital/AM Center visits,
# PM shift = Clinic/Poly Clinic visits. Office Work, Activities, Events,
# and Pharmacy rows are NOT calls even when they carry an AM/PM shift label
# (Office Work/Activities keep their own logged time-of-day for attendance
# tracking, which is a different thing from a doctor call).
AM_CALL_CATEGORIES = {"Hospital", "AM Center"}
PM_CALL_CATEGORIES = {"Clinic", "Poly Clinics"}


# ── helpers ──────────────────────────────────────────────────────────────────

def safe_str(v):
    if v is None or (isinstance(v, float) and v != v):  # NaN check
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def parse_members(raw):
    """Members stored as 'Name1, Name2' or just 'Name1' in Raw Data."""
    if not raw:
        return []
    return [n.strip() for n in str(raw).split(",") if n.strip() and n.strip().lower() != "nan"]


def minutes_between(t1, t2):
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
        try:
            a = datetime.strptime(t1.strip(), fmt)
            b = datetime.strptime(t2.strip(), fmt)
            return abs((b - a).total_seconds() / 60)
        except Exception:
            continue
    return 999


def fmt_hm(total_minutes):
    h = int(total_minutes // 60)
    m = int(total_minutes % 60)
    return f"{h}:{m:02d}"


# ── load & normalise raw data ─────────────────────────────────────────────────

def load_raw_data(paths):
    """Load Raw Data sheet(s) from one or more Excel files, return list of record dicts."""
    dfs = []
    for path in paths:
        xl = pd.ExcelFile(path)
        if "Raw Data" not in xl.sheet_names:
            print(f"  ⚠️  No 'Raw Data' sheet in {path} — skipping")
            continue
        df = pd.read_excel(path, sheet_name="Raw Data")
        dfs.append(df)
        print(f"  ✓ Loaded {len(df)} rows from {path}")

    if not dfs:
        sys.exit("❌  No Raw Data sheets found in any input file.")

    combined = pd.concat(dfs, ignore_index=True)
    print(f"  Total rows: {len(combined)}")

    records = []
    for _, row in combined.iterrows():
        r = {
            "team":               safe_str(row.get("Team", row.get("team"))),
            "user":               safe_str(row.get("user")),
            "territory":          safe_str(row.get("territory")),
            "date":               safe_str(row.get("date")),
            "time":               safe_str(row.get("time")),
            "acc_type_raw":       safe_str(row.get("acc_type_raw")),
            "acc_type_category":  safe_str(row.get("acc_type_category")),
            "shift":              safe_str(row.get("shift")),
            "visit_type_raw":     safe_str(row.get("visit_type_raw")),
            "visit_type_category": safe_str(row.get("visit_type_category")),
            "acc_id":             safe_str(row.get("acc_id")).rstrip(".0"),
            "acc_name":           safe_str(row.get("acc_name")),
            "doctor_key":         safe_str(row.get("doctor_key")).rstrip(".0"),
            "doctor_name":        safe_str(row.get("doctor_name")),
            "notes":              safe_str(row.get("notes")),
            "specialty":          safe_str(row.get("specialty")),
            "classification":     safe_str(row.get("classification")),
            "products":           safe_str(row.get("products")),
            # members: in Raw Data it's the user name only (single person per row)
            # but for coaching-day detection we need the original full list.
            # If the file was produced by a newer script version it already stores
            # only the user; older versions stored comma-separated lists.
            "members":            parse_members(row.get("members", "")),
            # Preserve any employee code already resolved in a prior run, so a
            # rebuild never silently drops a code that hierarchy name-matching
            # can no longer find (e.g. hierarchy file changed, name spelling
            # drifted). Used only as a fallback if a fresh hierarchy lookup
            # fails — a successful fresh lookup always wins.
            "existing_user_code": safe_str(
                row.get("user_code", row.get("Employee Code", ""))
            ).replace(".0", ""),
        }
        # Ensure members includes at least the user themselves
        if r["user"] and r["user"] not in r["members"]:
            r["members"] = [r["user"]] + r["members"]
        if r["user"]:
            records.append(r)

    return records


# ── de-duplication ────────────────────────────────────────────────────────────

def dedupe_exact_records(records, debug=False):
    """
    PulpoPlus exports 'Double' visit rows twice — for a manager doing a
    joint/coaching call, the same visit (same doctor, same time) shows up
    as two records credited to the same user, instead of one row per
    participant. That double-credits every downstream metric (Total
    Visits, PM/AM Calls, Product Calls, ...).

    A record is a duplicate here if it shares the same user, date, time,
    shift, and account/doctor — a real second visit by the same rep can
    never land on the exact same account at the exact same logged time on
    the same day. We deliberately do NOT require every field (like acc_id
    or notes) to match too, since duplicated export rows can carry a
    slightly different internal id or formatting even though they describe
    the same real visit — requiring an exact full-row match let some of
    these slip through uncaught.
    """
    seen = set()
    deduped = []
    dropped_by_user = Counter()
    for r in records:
        key = (
            r["user"], r["date"], r["time"], r["shift"],
            r["acc_id"] or r["doctor_key"] or r["acc_name"],
        )
        if key in seen:
            dropped_by_user[r["user"]] += 1
            continue
        seen.add(key)
        deduped.append(r)

    total_dropped = sum(dropped_by_user.values())
    if total_dropped:
        print(f"  ✓ Removed {total_dropped} duplicate record(s) "
              f"(Double-visit rows exported twice for the same user)")
        if debug:
            for user, n in dropped_by_user.most_common():
                print(f"     {user}: {n} duplicate(s) removed")
    return deduped


# ── territory backfill ────────────────────────────────────────────────────────

def backfill_territory(records, debug=False):
    user_date_territory = {}
    user_territory_counts = defaultdict(Counter)
    for r in records:
        t = r["territory"]
        if t and t.lower() not in ACC_TYPE_LABELS_LOWER:
            user_territory_counts[r["user"]][t] += 1
            key = (r["user"], r["date"])
            if key not in user_date_territory:
                user_date_territory[key] = t

    filled_same = filled_maj = unresolved = 0
    for r in records:
        if r["territory"] and r["territory"].lower() not in ACC_TYPE_LABELS_LOWER:
            continue
        r["territory"] = ""
        key = (r["user"], r["date"])
        if key in user_date_territory:
            r["territory"] = user_date_territory[key]
            filled_same += 1
        elif user_territory_counts[r["user"]]:
            r["territory"] = user_territory_counts[r["user"]].most_common(1)[0][0]
            filled_maj += 1
        else:
            unresolved += 1

    print(f"  ✓ Territory backfill: {filled_same} same-date, {filled_maj} majority, {unresolved} unresolved")


# ── manager identification ────────────────────────────────────────────────────

def identify_managers(records):
    territories_by_user = defaultdict(set)
    for r in records:
        t = r["territory"]
        if t and t.lower() not in ACC_TYPE_LABELS_LOWER:
            territories_by_user[r["user"]].add(t)
    return {u for u, terrs in territories_by_user.items() if len(terrs) > 1}


# ── coaching days ─────────────────────────────────────────────────────────────

COACHING_ACCOMPANIMENT_THRESHOLD = 0.8

def compute_coaching_days(records, managers, debug=False):
    """
    A (manager, rep, date) is a coaching day only if the rep worked BOTH
    an AM shift and a PM shift that day, AND the manager accompanied at
    least 80% of the rep's AM-shift call visits AND at least 80% of the
    rep's PM-shift call visits. A shift the rep didn't work that day does
    NOT get a free pass — if the rep has 0 visits in one shift, that day
    cannot qualify as coaching at all (it's just an ordinary double visit
    in whichever shift they did work, already reflected separately via
    Double Visit Days).

    "Manager accompanied a visit" = either the manager is listed in that
    row's Members, or the manager has their own row at the same account,
    date, and shift within COACHING_TIME_TOLERANCE_MINUTES of the rep's
    visit time.
    """
    # Manager's own call visits, indexed for the acc_id/time-proximity signal.
    mgr_visits_idx = defaultdict(list)  # (date, shift, acc_id) -> [(manager, time), ...]
    for r in records:
        if r["user"] not in managers or r["shift"] not in ("AM", "PM") or not r["date"] or not r["acc_id"]:
            continue
        mgr_visits_idx[(r["date"], r["shift"], r["acc_id"])].append((r["user"], r["time"]))

    # Reps' own call visits (real doctor calls only — same categories as
    # AM Calls/PM Calls), grouped so we can compute per-shift totals.
    rep_visits = defaultdict(list)  # (rep, date, shift) -> [record, ...]
    for r in records:
        if r["user"] in managers or r["shift"] not in ("AM", "PM") or not r["date"]:
            continue
        if r["shift"] == "AM" and r["acc_type_category"] not in AM_CALL_CATEGORIES:
            continue
        if r["shift"] == "PM" and r["acc_type_category"] not in PM_CALL_CATEGORIES:
            continue
        rep_visits[(r["user"], r["date"], r["shift"])].append(r)

    # accompanied[(manager, rep, date)][shift] = count of rep visits that shift
    # the manager was present for.
    accompanied = defaultdict(lambda: {"AM": 0, "PM": 0})
    for (rep, date, shift), visits in rep_visits.items():
        for r in visits:
            matched_managers = {m for m in r["members"] if m in managers and m != rep}
            if r["acc_id"]:
                for m_user, m_time in mgr_visits_idx.get((date, shift, r["acc_id"]), []):
                    if not m_time or not r["time"] or minutes_between(m_time, r["time"]) <= COACHING_TIME_TOLERANCE_MINUTES:
                        matched_managers.add(m_user)
            for m in matched_managers:
                accompanied[(m, rep, date)][shift] += 1

    rep_shift_totals = {key: len(visits) for key, visits in rep_visits.items()}
    team_by_user = team_by_user_map(records)

    coaching_days_by_manager = defaultdict(int)
    detail_rows = []
    partial = []
    for (manager, rep, date), acc in accompanied.items():
        am_total = rep_shift_totals.get((rep, date, "AM"), 0)
        pm_total = rep_shift_totals.get((rep, date, "PM"), 0)
        am_acc, pm_acc = acc["AM"], acc["PM"]
        am_ratio = (am_acc / am_total) if am_total else None
        pm_ratio = (pm_acc / pm_total) if pm_total else None
        am_ok = am_ratio is not None and am_ratio >= COACHING_ACCOMPANIMENT_THRESHOLD
        pm_ok = pm_ratio is not None and pm_ratio >= COACHING_ACCOMPANIMENT_THRESHOLD

        # A full coaching day requires the rep to have worked BOTH shifts
        # that day, with the manager accompanying at least 80% of each.
        # A rep with 0 visits in one shift can no longer get a free pass —
        # that's just a single-shift double visit, not a coaching day.
        is_coaching = am_total > 0 and pm_total > 0 and am_ok and pm_ok

        if is_coaching:
            coaching_days_by_manager[manager] += 1
            detail_rows.append({
                "Team": team_by_user.get(manager, ""),
                "Manager": manager, "Rep": rep, "Date": date,
                "AM Visits": am_total, "AM Accompanied": am_acc,
                "AM %": round(am_ratio * 100, 0) if am_ratio is not None else "",
                "PM Visits": pm_total, "PM Accompanied": pm_acc,
                "PM %": round(pm_ratio * 100, 0) if pm_ratio is not None else "",
            })
        else:
            partial.append((manager, rep, date, am_total, am_acc, am_ratio, pm_total, pm_acc, pm_ratio))

    if debug and partial:
        print(f"\n  🔍 Overlaps below {int(COACHING_ACCOMPANIMENT_THRESHOLD*100)}% threshold "
              f"(counted as normal double visits, not coaching):")
        for mgr, rep, date, am_t, am_a, am_r, pm_t, pm_a, pm_r in sorted(partial):
            am_s = f"{am_a}/{am_t} ({am_r:.0%})" if am_t else "—"
            pm_s = f"{pm_a}/{pm_t} ({pm_r:.0%})" if pm_t else "—"
            print(f"     {mgr} + {rep} on {date}: AM {am_s}, PM {pm_s}")

    detail_rows.sort(key=lambda x: (x["Manager"], x["Date"], x["Rep"]))
    return dict(coaching_days_by_manager), pd.DataFrame(detail_rows)


# ── main summary ──────────────────────────────────────────────────────────────

def _compute_user_row(user, user_records, team_by_user, managers,
                       coaching_days_by_manager, coaching_dates_by_manager,
                       single_date=None):
    """
    Compute one summary row for `user` from `user_records`.

    If `single_date` is given, `user_records` is expected to already be
    filtered to just that date, and Coaching Days is computed by checking
    whether this specific date is one of the manager's coaching dates
    (rather than using the whole-period total). Every other metric is
    recomputed fresh from the given records — so when called once per day,
    each day's numbers are genuinely that day's numbers, not a slice of a
    pre-averaged period total.
    """
    # Territory: prefer the hierarchy's official assignment when
    # available (it's authoritative), falling back to territories
    # inferred from the rep's own visited accounts otherwise.
    hierarchy_territories = sorted(set(
        r.get("hierarchy_territory") for r in user_records
        if r.get("hierarchy_territory")
    ))
    if hierarchy_territories:
        territory_str = "; ".join(hierarchy_territories)
    else:
        territories = sorted(set(
            r["territory"] for r in user_records
            if r["territory"] and r["territory"].lower() not in ACC_TYPE_LABELS_LOWER
        ))
        territory_str = "; ".join(territories)

    # Shift splits
    am_records = [r for r in user_records if r["shift"] == "AM"]
    pm_records = [r for r in user_records if r["shift"] == "PM"]
    ph_records = [r for r in user_records if r["acc_type_category"] == "Pharmacy"]

    # "Real" AM/PM visits: AM = Hospital/AM Center, PM = Clinic/Poly
    # Clinics — this is what defines both Shift Days and Calls. Office
    # Work/Activities (which carry their own logged AM/PM time) and
    # Pharmacy (no doctor shift) don't count toward either.
    am_call_records = [r for r in am_records if r["acc_type_category"] in AM_CALL_CATEGORIES]
    pm_call_records = [r for r in pm_records if r["acc_type_category"] in PM_CALL_CATEGORIES]

    # Working days
    dates_with_visits = sorted(set(r["date"] for r in user_records if r["date"]))
    working_days = len(dates_with_visits)

    am_shift_dates = set(r["date"] for r in am_call_records if r["date"])
    pm_shift_dates = set(r["date"] for r in pm_call_records if r["date"])
    # Complete Field Days reflects actual doctor-call coverage only —
    # Office Work no longer props up the AM slot here.
    am_shift_days  = len(am_shift_dates)
    pm_shift_days  = len(pm_shift_dates)
    total_shifts   = len(am_shift_dates) + len(pm_shift_dates)
    complete_field_days = round(total_shifts / 2, 1)

    double_visit_dates = set(
        r["date"] for r in user_records
        if r["date"] and r["visit_type_category"] == "Double"
    )
    double_visit_days = len(double_visit_dates)

    if single_date is not None:
        # Daily row: was THIS specific date a coaching day for this manager?
        coaching_days = 1 if single_date in coaching_dates_by_manager.get(user, set()) else 0
    else:
        # Whole-period row (original behaviour): total coaching day count.
        coaching_days = coaching_days_by_manager.get(user, 0)

    # Doctor coverage
    doctors_by_category = defaultdict(set)
    for r in am_records + pm_records:
        if r["doctor_key"] and r["acc_type_category"] in ("Clinic", "Poly Clinics", "Hospital", "AM Center"):
            doctors_by_category[r["acc_type_category"]].add(r["doctor_key"])

    am_unique_doctors  = set(r["doctor_key"] for r in am_records if r["doctor_key"])
    pm_unique_doctors  = set(r["doctor_key"] for r in pm_records if r["doctor_key"])
    am_specialties     = set(r["specialty"] for r in am_records if r["specialty"])
    pm_specialties     = set(r["specialty"] for r in pm_records if r["specialty"])

    # "Calls" = actual doctor visits only (Hospital/AM Center for AM,
    # Clinic/Poly Clinics for PM) — excludes Office Work/Activities rows
    # that merely happen to carry an AM/PM shift label.
    am_calls = len(am_call_records)
    pm_calls = len(pm_call_records)
    pm_call_rate = round(pm_calls / pm_shift_days, 1) if pm_shift_days else 0.0
    am_call_rate = round(am_calls / am_shift_days, 1) if am_shift_days else 0.0

    # Pharmacy
    pharmacies_visited = len(ph_records)
    pharmacy_unique = set()
    for r in ph_records:
        key = r["acc_id"] if r["acc_id"] else r["acc_name"]
        if key:
            pharmacy_unique.add(key)
    pharmacies_covered = len(pharmacy_unique)

    # Shift durations — also keep the RAW total minutes (not just the
    # formatted average), so the app can correctly re-average across any
    # date range the user picks later (summing raw minutes / summing raw
    # shift-days), instead of averaging pre-computed daily averages.
    def shift_duration_minutes(shift_recs):
        by_day = defaultdict(list)
        for r in shift_recs:
            if r["date"] and r["time"]:
                by_day[r["date"]].append(r["time"])
        durations = []
        for _, times in by_day.items():
            times_sorted = sorted(times)
            if len(times_sorted) >= 2:
                durations.append(minutes_between(times_sorted[0], times_sorted[-1]))
            else:
                durations.append(0.0)
        return durations  # one value per day present in this slice

    am_durations = shift_duration_minutes(am_records)
    pm_durations = shift_duration_minutes(pm_records)
    am_total_dur_min = sum(am_durations)
    pm_total_dur_min = sum(pm_durations)
    am_avg_dur = (am_total_dur_min / len(am_durations)) if am_durations else 0.0
    pm_avg_dur = (pm_total_dur_min / len(pm_durations)) if pm_durations else 0.0
    # Overall = straight sum of the AM and PM averages, not a separate
    # recomputation over the combined AM+PM records.
    overall_avg = am_avg_dur + pm_avg_dur

    # AM average start time
    am_start_times = []
    for date_key in am_shift_dates:
        day_times = sorted(r["time"] for r in am_records if r["date"] == date_key and r["time"])
        if day_times:
            am_start_times.append(day_times[0])
    am_avg_start_time = ""
    am_start_total_min, am_start_count = 0, 0
    if am_start_times:
        for t in am_start_times:
            for fmt in ("%H:%M:%S", "%H:%M"):
                try:
                    dt = datetime.strptime(t.strip(), fmt)
                    am_start_total_min += dt.hour * 60 + dt.minute
                    am_start_count += 1
                    break
                except Exception:
                    continue
        if am_start_count:
            avg_min = am_start_total_min / am_start_count
            am_avg_start_time = f"{int(avg_min // 60):02d}:{int(avg_min % 60):02d}"

    # Product calls
    product_call_counts = defaultdict(int)
    for r in user_records:
        if r["shift"] not in ("AM", "PM") or not r["products"]:
            continue
        for prod in r["products"].split(","):
            prod = prod.strip()
            if prod:
                product_call_counts[prod] += 1
    products_summary = ", ".join(
        f"{p}({c})" for p, c in sorted(product_call_counts.items(), key=lambda x: -x[1])
    )
    total_product_calls  = sum(product_call_counts.values())
    distinct_products    = len(product_call_counts)

    # Activities, Events & Office Work: count by DAY, not by row. If a
    # user has an entry in both AM and PM on the same date, that's one
    # continuous full day — count it as 1. If only one shift has an
    # entry that date, it's a half day — count it as 0.5. This applies
    # per date, then sums across all the user's dates for that category.
    def day_based_count(category):
        by_date_shifts = defaultdict(set)
        for r in user_records:
            if r["acc_type_category"] == category and r["date"]:
                by_date_shifts[r["date"]].add(r["shift"])
        total = 0.0
        for shifts in by_date_shifts.values():
            if "AM" in shifts and "PM" in shifts:
                total += 1
            elif "AM" in shifts or "PM" in shifts:
                total += 0.5
        return total

    no_activities     = day_based_count("Activities")
    no_events         = day_based_count("Events")
    office_work_days  = day_based_count("Office Work")

    row = {
        "Team":                       team_by_user.get(user, ""),
        "User":                       user,
        "Territory":                  territory_str,
        "Is Manager":                 "Yes" if user in managers else "No",
        "Working Days":               working_days,
        "Complete Field Days":        complete_field_days,
        "Office Work Days":           office_work_days,
        "No. of Activities":          no_activities,
        "No. of Events":              no_events,
        "Double Visit Days":          double_visit_days,
        "Coaching Days":              coaching_days,
        "AM Shift Days":              am_shift_days,
        "PM Shift Days":              pm_shift_days,
        "Total PM Covered":           (len(doctors_by_category.get("Clinic", set()))
                                       + len(doctors_by_category.get("Poly Clinics", set()))),
        "PM Calls":                   pm_calls,
        "PM Call Rate":               pm_call_rate,
        "Clinic Doctors Covered":     len(doctors_by_category.get("Clinic", set())),
        "Poly Clinic Doctors Covered": len(doctors_by_category.get("Poly Clinics", set())),
        "PM Unique Doctors":          len(pm_unique_doctors),
        "PM Specialties Covered":     len(pm_specialties),
        "Avg PM Shift Duration (h:mm)": fmt_hm(pm_avg_dur),
        "PM Shift Duration Total Min": round(pm_total_dur_min, 1),
        "Total AM Covered":           (len(doctors_by_category.get("AM Center", set()))
                                       + len(doctors_by_category.get("Hospital", set()))),
        "AM Calls":                   am_calls,
        "AM Call Rate":               am_call_rate,
        "AM Center Doctors Covered":  len(doctors_by_category.get("AM Center", set())),
        "Hospital Doctors Covered":   len(doctors_by_category.get("Hospital", set())),
        "Avg AM Starting Time":       am_avg_start_time,
        "AM Start Time Total Min":    am_start_total_min,
        "AM Start Time Count":        am_start_count,
        "AM Unique Doctors":          len(am_unique_doctors),
        "AM Specialties Covered":     len(am_specialties),
        "Avg AM Shift Duration (h:mm)": fmt_hm(am_avg_dur),
        "AM Shift Duration Total Min": round(am_total_dur_min, 1),
        "Pharmacies Visited":         pharmacies_visited,
        "Pharmacies Covered":         pharmacies_covered,
        "Total Product Calls":        total_product_calls,
        "Distinct Products Promoted": distinct_products,
        "Avg Field Duration Overall (h:mm)": fmt_hm(overall_avg),
        "Product Detail":             products_summary,
        "Total Visits":               am_calls + pm_calls,
    }
    if single_date is not None:
        row = {"Date": single_date, **row}
    return row


def compute_summary(records, managers=None, coaching_days_by_manager=None,
                    coaching_detail_df=None, debug=False, group_by_date=False):
    """
    Build the Summary sheet.

    group_by_date=False (default): one row per user, aggregated across
    every record passed in — this is the original whole-period behaviour,
    unchanged, for callers that still want a single period total.

    group_by_date=True: one row per (user, date) — every metric is
    recomputed fresh from that single day's records, so each daily row is
    a true daily value rather than a slice of a period-wide average.
    """
    managers = managers or set()
    coaching_days_by_manager = coaching_days_by_manager or {}

    # Build coaching date sets per manager for per-day classification
    coaching_dates_by_manager = defaultdict(set)
    if coaching_detail_df is not None and not coaching_detail_df.empty:
        for _, row in coaching_detail_df.iterrows():
            coaching_dates_by_manager[row["Manager"]].add(str(row["Date"]))

    team_by_user = team_by_user_map(records)
    summary_rows = []

    if not group_by_date:
        by_user = defaultdict(list)
        for r in records:
            by_user[r["user"]].append(r)
        for user, user_records in sorted(by_user.items()):
            summary_rows.append(_compute_user_row(
                user, user_records, team_by_user, managers,
                coaching_days_by_manager, coaching_dates_by_manager,
                single_date=None,
            ))
    else:
        by_user_date = defaultdict(list)
        for r in records:
            if r["date"]:
                by_user_date[(r["user"], r["date"])].append(r)
        for (user, date), day_records in sorted(by_user_date.items()):
            summary_rows.append(_compute_user_row(
                user, day_records, team_by_user, managers,
                coaching_days_by_manager, coaching_dates_by_manager,
                single_date=date,
            ))

    return pd.DataFrame(summary_rows)


# ── timing report (last visit of the day) ──────────────────────────────────

def compute_timing_report(records, early_cutoff="15:00", late_cutoff="18:00"):
    """
    For every (user, date), find the LAST logged visit time that day and
    flag it:
      • "Early"  if the last visit was before `early_cutoff` (default 3pm)
      • "Late"   if the last visit was at/after `late_cutoff` (default 6pm)
      • ""       (no flag) otherwise — the normal/expected range

    Only field-visit records with a real time are considered (Office Work/
    Activities/Events rows are excluded, same categories treated as non-
    field elsewhere in this script).
    """
    def _parse_minutes(t):
        if not t:
            return None
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                dt = datetime.strptime(t.strip(), fmt)
                return dt.hour * 60 + dt.minute
            except Exception:
                continue
        return None

    early_cutoff_min = _parse_minutes(early_cutoff)
    late_cutoff_min = _parse_minutes(late_cutoff)

    FIELD_CATEGORIES = set(AM_CALL_CATEGORIES) | set(PM_CALL_CATEGORIES)

    team_by_user = team_by_user_map(records)
    by_user_date = defaultdict(list)
    for r in records:
        if r["date"] and r["time"] and r["acc_type_category"] in FIELD_CATEGORIES:
            by_user_date[(r["user"], r["date"])].append(r)

    rows = []
    for (user, date), day_records in sorted(by_user_date.items()):
        times_min = [(t, m) for t in (r["time"] for r in day_records)
                     for m in [_parse_minutes(t)] if m is not None]
        if not times_min:
            continue
        last_time_str, last_time_min = max(times_min, key=lambda tm: tm[1])

        flag = ""
        if early_cutoff_min is not None and last_time_min < early_cutoff_min:
            flag = "Early"
        elif late_cutoff_min is not None and last_time_min >= late_cutoff_min:
            flag = "Late"

        rows.append({
            "Team":            team_by_user.get(user, ""),
            "User":            user,
            "Date":            date,
            "Last Visit Time": last_time_str,
            "Flag":            flag,
        })

    return pd.DataFrame(rows)


# ── shared team lookup ────────────────────────────────────────────────────────

def team_by_user_map(records):
    """
    All distinct `team` values seen for each user, joined together (e.g.
    "PLATINUM 1; PLATINUM 2"). A single majority pick breaks for people who
    genuinely belong to more than one team — e.g. a manager overseeing two
    teams — so instead of guessing which one is "right", every team that
    genuinely appears for that user is shown.
    """
    teams_by_user = defaultdict(set)
    for r in records:
        if r["team"]:
            teams_by_user[r["user"]].add(r["team"])
    return {u: "; ".join(sorted(teams)) for u, teams in teams_by_user.items()}


# ── specialty × classification ────────────────────────────────────────────────

def compute_specialty_classification(records, group_by_date=False):
    team_by_user = team_by_user_map(records)
    groups = defaultdict(lambda: {"calls": 0, "doctors": set()})
    for r in records:
        if r["shift"] not in ("AM", "PM"):
            continue
        specialty = r["specialty"] or "Unknown"
        classification = r["classification"] or "Unknown"
        if group_by_date:
            key = (r["user"], r["date"], specialty, classification, r["shift"])
        else:
            key = (r["user"], specialty, classification, r["shift"])
        groups[key]["calls"] += 1
        if r["doctor_key"]:
            groups[key]["doctors"].add(r["doctor_key"])

    rows = []
    for key, agg in sorted(groups.items(), key=lambda kv: tuple(str(x) for x in kv[0])):
        if group_by_date:
            user, date, spec, cls, shift = key
        else:
            user, spec, cls, shift = key
            date = None
        row = {
            "Team":            team_by_user.get(user, ""),
            "User":            user,
            "Specialty":       spec,
            "Classification":  cls,
            "Shift":           shift,
            "Call Count":      agg["calls"],
            "Unique Doctors":  len(agg["doctors"]),
        }
        if group_by_date:
            row = {"Date": date, **row}
        rows.append(row)
    return pd.DataFrame(rows)


# ── product calls per specialty ───────────────────────────────────────────────

def compute_product_calls(records, group_by_date=False):
    team_by_user = team_by_user_map(records)
    groups = defaultdict(lambda: {"calls": 0, "doctors": set()})
    for r in records:
        if r["shift"] not in ("AM", "PM") or not r["products"]:
            continue
        specialty = r["specialty"] or "Unknown"
        for prod in r["products"].split(","):
            prod = prod.strip()
            if not prod:
                continue
            if group_by_date:
                key = (r["user"], r["date"], prod, r["shift"], specialty)
            else:
                key = (r["user"], prod, r["shift"], specialty)
            groups[key]["calls"] += 1
            if r["doctor_key"]:
                groups[key]["doctors"].add(r["doctor_key"])

    rows = []
    for key, agg in sorted(groups.items(), key=lambda kv: tuple(str(x) for x in kv[0])):
        if group_by_date:
            user, date, product, shift, specialty = key
        else:
            user, product, shift, specialty = key
            date = None
        row = {
            "Team":           team_by_user.get(user, ""),
            "User":           user,
            "Product":        product,
            "Shift":          shift,
            "Specialty":      specialty,
            "Call Count":     agg["calls"],
            "Unique Doctors": len(agg["doctors"]),
        }
        if group_by_date:
            row = {"Date": date, **row}
        rows.append(row)
    return pd.DataFrame(rows)


# ── Excel styling ─────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")
BORDER_SIDE   = Side(style="thin", color="BDD7EE")
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                       top=BORDER_SIDE,  bottom=BORDER_SIDE)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def style_worksheet(ws, num_data_rows, num_cols, freeze_cell="A2"):
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, num_data_rows + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = CELL_BORDER
            if row_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER
            else:
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL
                cell.alignment = LEFT
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)
    ws.freeze_panes = freeze_cell


# ── save ──────────────────────────────────────────────────────────────────────

def _inject_user_code(df: "pd.DataFrame", hmap, user_col: str = "User") -> "pd.DataFrame":
    """
    Insert an 'Employee Code' column immediately after `user_col`.
    Resolves codes via the hierarchy map; leaves blank when not found.
    Works on a copy — does not mutate the original DataFrame.
    """
    if hmap is None or user_col not in df.columns:
        return df
    df = df.copy()
    codes = []
    for name in df[user_col]:
        emp = hmap.lookup(name=str(name)) if name else None
        codes.append(emp.code if emp and emp.code else "")
    pos = df.columns.get_loc(user_col) + 1
    df.insert(pos, "Employee Code", codes)
    return df


def save_output(summary_df, records, output_path,
                specialty_class_df=None, product_df=None,
                coaching_detail_df=None, hmap=None, timing_df=None):
    print(f"\n💾 Saving to {output_path} ...")

    # ── inject Employee Code into every analytical sheet ─────────────────────
    s_df   = _inject_user_code(summary_df,         hmap, user_col="User")
    sp_df  = _inject_user_code(specialty_class_df, hmap, user_col="User") \
             if specialty_class_df is not None else specialty_class_df
    pr_df  = _inject_user_code(product_df,         hmap, user_col="User") \
             if product_df is not None else product_df

    # Coaching Days has two user columns: Manager and Rep — add code for both
    cd_df = coaching_detail_df
    if cd_df is not None and not cd_df.empty and hmap is not None:
        cd_df = cd_df.copy()
        mgr_codes = [
            (hmap.lookup(name=str(n)) or type("", (), {"code": ""})()).code
            if n else ""
            for n in cd_df["Manager"]
        ]
        rep_codes = [
            (hmap.lookup(name=str(n)) or type("", (), {"code": ""})()).code
            if n else ""
            for n in cd_df["Rep"]
        ]
        cd_df.insert(cd_df.columns.get_loc("Manager") + 1, "Manager Code", mgr_codes)
        cd_df.insert(cd_df.columns.get_loc("Rep")     + 1, "Rep Code",     rep_codes)

    # Raw Data — use acc_id (already in the data) as the doctor/account code,
    # and add User Code next to the user column
    raw_df = pd.DataFrame(records)
    raw_df["members"] = raw_df["user"]
    raw_df = raw_df.rename(columns={"team": "Team"})
    raw_df.insert(0, "Team", raw_df.pop("Team"))

    existing_codes = raw_df.pop("existing_user_code") if "existing_user_code" in raw_df.columns else None

    if "user" in raw_df.columns:
        user_codes = []
        for i, n in enumerate(raw_df["user"]):
            emp = hmap.lookup(name=str(n)) if (hmap is not None and n) else None
            if emp and emp.code:
                user_codes.append(emp.code)
            elif existing_codes is not None:
                # Fresh hierarchy lookup failed (or no hierarchy loaded) —
                # fall back to whatever code this record already had.
                user_codes.append(existing_codes.iloc[i])
            else:
                user_codes.append("")
        raw_df.insert(raw_df.columns.get_loc("user") + 1, "user_code", user_codes)

    def _format_date_col(df, col_name):
        if df is None or df.empty or col_name not in df.columns:
            return df
        df = df.copy()
        def _fmt(val):
            if val is None or pd.isna(val) or not str(val).strip():
                return ""
            s = str(val).strip()
            try:
                dt = datetime.strptime(s[:10], "%Y-%m-%d")
                return dt.strftime("%b %d, %Y")
            except ValueError:
                pass
            return s
        df[col_name] = df[col_name].apply(_fmt)
        return df

    s_df = _format_date_col(s_df, "Date")
    cd_df = _format_date_col(cd_df, "Date")
    if cd_df is not None and "Coaching Date" in cd_df.columns:
        cd_df = _format_date_col(cd_df, "Coaching Date")
    timing_df_fmt = _format_date_col(timing_df, "Date") if timing_df is not None else None
    raw_df_fmt = _format_date_col(raw_df, "date") if not raw_df.empty else raw_df
    if raw_df_fmt is not None and "Date" in raw_df_fmt.columns:
        raw_df_fmt = _format_date_col(raw_df_fmt, "Date")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        s_df.to_excel(writer, sheet_name="Summary", index=False)
        style_worksheet(writer.sheets["Summary"], len(s_df), len(s_df.columns), freeze_cell="D2")
        print(f"   ✓ Summary:            {len(s_df)} users")

        if sp_df is not None and not sp_df.empty:
            sp_df.to_excel(writer, sheet_name="Specialty x Class", index=False)
            style_worksheet(writer.sheets["Specialty x Class"], len(sp_df), len(sp_df.columns))
            print(f"   ✓ Specialty x Class:  {len(sp_df)} rows")

        if pr_df is not None and not pr_df.empty:
            pr_df.to_excel(writer, sheet_name="Product Calls per spec", index=False)
            style_worksheet(writer.sheets["Product Calls per spec"], len(pr_df), len(pr_df.columns))
            print(f"   ✓ Product Calls:      {len(pr_df)} rows")

        if cd_df is not None and not cd_df.empty:
            cd_df.to_excel(writer, sheet_name="Coaching Days", index=False)
            style_worksheet(writer.sheets["Coaching Days"], len(cd_df), len(cd_df.columns))
            print(f"   ✓ Coaching Days:      {len(cd_df)} rows")

        if timing_df_fmt is not None and not timing_df_fmt.empty:
            t_df = _inject_user_code(timing_df_fmt, hmap, user_col="User")
            t_df.to_excel(writer, sheet_name="Timing", index=False)
            style_worksheet(writer.sheets["Timing"], len(t_df), len(t_df.columns))
            print(f"   ✓ Timing:             {len(t_df)} rows")

        if raw_df_fmt is not None and not raw_df_fmt.empty:
            raw_df_fmt.to_excel(writer, sheet_name="Raw Data", index=False)
            style_worksheet(writer.sheets["Raw Data"], len(raw_df), len(raw_df.columns))
            print(f"   ✓ Raw Data:           {len(raw_df)} records")

    print(f"\n✅ Done → {output_path}")


# ── entry point ───────────────────────────────────────────────────────────────

def auto_output_filename(records, fallback_base="output"):
    """
    Build an output filename like "June 2026 - Platinum 1.xlsx" from the
    raw data itself: the month is whichever calendar month appears most
    often across all records' dates, and the team is whichever team has
    the most records (the "heavily existed"/dominant team). Falls back to
    a generic name if dates or teams can't be determined at all.
    """
    month_counter = Counter()
    for r in records:
        d = r.get("date")
        if not d:
            continue
        try:
            dt = datetime.strptime(str(d)[:10], "%Y-%m-%d")
            month_counter[(dt.year, dt.month)] += 1
        except ValueError:
            continue

    team_counter = Counter(r["team"] for r in records if r.get("team"))

    if month_counter and team_counter:
        (year, month), _ = month_counter.most_common(1)[0]
        month_name = datetime(year, month, 1).strftime("%B")
        dominant_team = team_counter.most_common(1)[0][0]
        # Title-case for a cleaner filename (PLATINUM 1 → Platinum 1)
        team_display = dominant_team.title()
        # Strip characters that aren't safe in Windows filenames
        team_display = re.sub(r'[\\/:*?"<>|]', "", team_display)
        return f"{month_name} {year} - {team_display}.xlsx"

    return f"{fallback_base}_rebuilt.xlsx"


def main():
    parser = argparse.ArgumentParser(
        description="Rebuild summary sheets from Raw Data in a pulpoplus Excel file"
    )
    parser.add_argument("inputs", nargs="+", help="One or more input Excel files with a 'Raw Data' sheet")
    parser.add_argument("--out",       default=None, help="Output file path (default: rebuilt_<first_input>)")
    parser.add_argument("--team",      default=None, help="Team name to fill in for any record missing one")
    parser.add_argument("--hierarchy", default=None,
                        help="Path to the 2026 Team Structure .xlsx (auto-discovered if omitted)")
    parser.add_argument("--no-hierarchy", action="store_true",
                        help="Skip hierarchy file even if one is found")
    parser.add_argument("--debug", action="store_true", help="Print extra diagnostic info")
    args = parser.parse_args()

    out_path = args.out  # resolved after records load if not explicitly given

    print("\n" + "="*65)
    print("PulpoPlus Summary Rebuilder")
    print("="*65)
    print(f"\n📂 Input(s):  {', '.join(args.inputs)}")
    if out_path:
        print(f"📄 Output:    {out_path}\n")
    else:
        print(f"📄 Output:    (auto-named from data — determined after loading)\n")

    # ── hierarchy ─────────────────────────────────────────────────────────────
    hmap = None
    if not args.no_hierarchy:
        print("🏢 Loading 2026 team hierarchy...")
        hmap = _load_hierarchy_if_available(args.hierarchy)
        if hmap is None:
            print("   (no hierarchy file found — team names will come from Raw Data only)")

    print("📥 Loading Raw Data...")
    records = load_raw_data(args.inputs)
    print(f"   Total records loaded: {len(records)}")

    # Enrich team from hierarchy before the manual --team override
    if hmap is not None:
        print("\n🔗 Enriching records with hierarchy team assignments...")
        enriched = hmap.enrich_records(records, warn=args.debug)
        print(f"   ✓ {enriched} record(s) assigned a team from hierarchy")

    if args.team:
        filled = 0
        for r in records:
            if not r["team"]:
                r["team"] = args.team
                filled += 1
        if filled:
            print(f"   ✓ Filled Team = '{args.team}' for {filled} record(s) still missing one")

    if not out_path:
        base = args.inputs[0].replace(".xlsx", "").replace(".xls", "")
        out_path = auto_output_filename(records, fallback_base=base)
        print(f"   📄 Auto-named output: {out_path}")

    print("\n🧹 Removing exact-duplicate records...")
    records = dedupe_exact_records(records, debug=args.debug)
    print(f"   Records after dedup: {len(records)}")

    print("\n🌍 Backfilling territory...")
    backfill_territory(records, debug=args.debug)

    print("\n👔 Identifying managers...")
    managers = identify_managers(records)
    print(f"   Managers found ({len(managers)}): {sorted(managers)}")

    print("\n🤝 Computing coaching days...")
    coaching_days_by_manager, coaching_detail_df = compute_coaching_days(
        records, managers, debug=args.debug
    )
    for mgr, days in sorted(coaching_days_by_manager.items()):
        print(f"   {mgr}: {days} coaching day(s)")

    print("\n📊 Computing per-user-per-day summary...")
    summary_df = compute_summary(
        records,
        managers=managers,
        coaching_days_by_manager=coaching_days_by_manager,
        coaching_detail_df=coaching_detail_df,
        debug=args.debug,
        group_by_date=True,
    )

    print("\n🕐 Computing timing report (last visit per day)...")
    timing_df = compute_timing_report(records)

    print("\n🔬 Computing specialty × classification...")
    spec_df = compute_specialty_classification(records)

    print("\n💊 Computing product calls per specialty...")
    prod_df = compute_product_calls(records)

    save_output(
        summary_df, records, out_path,
        specialty_class_df=spec_df,
        product_df=prod_df,
        coaching_detail_df=coaching_detail_df,
        hmap=hmap,
        timing_df=timing_df,
    )


if __name__ == "__main__":
    main()